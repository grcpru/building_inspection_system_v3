"""
Report Utilities
================
Shared utilities for report generation across all report types.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import pandas as pd


class ReportStyler:
    """Shared styling utilities for Excel reports"""
    
    # Color scheme
    HEADER_COLOR = "366092"
    ACCENT_COLOR = "4472C4"
    SUCCESS_COLOR = "70AD47"
    WARNING_COLOR = "FFC000"
    DANGER_COLOR = "C00000"
    LIGHT_GRAY = "F2F2F2"
    
    @staticmethod
    def style_header_row(worksheet, row=1):
        """Apply header styling to a row"""
        header_fill = PatternFill(
            start_color=ReportStyler.HEADER_COLOR,
            end_color=ReportStyler.HEADER_COLOR,
            fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
    
    @staticmethod
    def auto_adjust_column_width(worksheet):
        """Auto-adjust column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def add_border_to_range(worksheet, start_row, end_row, start_col, end_col):
        """Add borders to a range of cells"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                worksheet.cell(row=row, column=col).border = thin_border
    
    @staticmethod
    def highlight_row_by_condition(worksheet, row, condition_value, colors_map):
        """
        Highlight a row based on condition
        
        Args:
            worksheet: Excel worksheet
            row: Row number
            condition_value: Value to check
            colors_map: Dict mapping values to colors
        """
        if condition_value in colors_map:
            fill = PatternFill(
                start_color=colors_map[condition_value],
                end_color=colors_map[condition_value],
                fill_type="solid"
            )
            for cell in worksheet[row]:
                cell.fill = fill


class ReportDataProcessor:
    """Shared data processing utilities"""
    
    @staticmethod
    def calculate_completion_rate(completed, total):
        """Calculate percentage completion rate"""
        if total == 0:
            return 0.0
        return round((completed / total) * 100, 1)
    
    @staticmethod
    def format_date(date_value):
        """Format date consistently"""
        if pd.isna(date_value):
            return "N/A"
        
        if isinstance(date_value, str):
            try:
                date_obj = pd.to_datetime(date_value)
                return date_obj.strftime('%Y-%m-%d')
            except:
                return date_value
        
        return date_value.strftime('%Y-%m-%d') if hasattr(date_value, 'strftime') else str(date_value)
    
    @staticmethod
    def get_priority_color(priority):
        """Get color based on priority level"""
        priority_colors = {
            'urgent': ReportStyler.DANGER_COLOR,
            'high': ReportStyler.WARNING_COLOR,
            'medium': ReportStyler.ACCENT_COLOR,
            'normal': ReportStyler.SUCCESS_COLOR,
            'low': ReportStyler.LIGHT_GRAY
        }
        return priority_colors.get(str(priority).lower(), ReportStyler.LIGHT_GRAY)
    
    @staticmethod
    def get_status_color(status):
        """Get color based on status"""
        status_colors = {
            'completed': ReportStyler.SUCCESS_COLOR,
            'in_progress': ReportStyler.WARNING_COLOR,
            'in progress': ReportStyler.WARNING_COLOR,
            'open': ReportStyler.DANGER_COLOR,
            'pending': ReportStyler.ACCENT_COLOR
        }
        return status_colors.get(str(status).lower(), ReportStyler.LIGHT_GRAY)
    
    @staticmethod
    def safe_division(numerator, denominator, default=0):
        """Safely divide two numbers"""
        try:
            if denominator == 0:
                return default
            return round(numerator / denominator, 2)
        except:
            return default


class ReportMetadata:
    """Generate report metadata"""
    
    @staticmethod
    def create_metadata_sheet(workbook, report_info):
        """
        Create a metadata sheet with report information
        
        Args:
            workbook: Excel workbook
            report_info: Dict with report metadata
        """
        ws = workbook.create_sheet("Report Info", 0)
        
        # Report information
        metadata = [
            ("Report Title", report_info.get('title', 'Report')),
            ("Generated Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("Generated By", report_info.get('generated_by', 'System')),
            ("Report Type", report_info.get('report_type', 'N/A')),
            ("Data Period", report_info.get('period', 'N/A')),
            ("", ""),
            ("Summary Statistics", ""),
            ("Total Records", report_info.get('total_records', 0)),
            ("Filters Applied", report_info.get('filters', 'None')),
        ]
        
        # Write metadata
        for idx, (key, value) in enumerate(metadata, start=1):
            ws.cell(row=idx, column=1, value=key).font = Font(bold=True)
            ws.cell(row=idx, column=2, value=value)
        
        # Style
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        
        return ws


def create_summary_dataframe(metrics_dict):
    """
    Convert metrics dictionary to a formatted summary DataFrame
    
    Args:
        metrics_dict: Dict with metric names and values
    
    Returns:
        pd.DataFrame: Formatted summary
    """
    summary_data = []
    
    for metric_name, metric_value in metrics_dict.items():
        # Format metric name (replace underscores, capitalize)
        display_name = metric_name.replace('_', ' ').title()
        
        # Format value
        if isinstance(metric_value, float):
            formatted_value = f"{metric_value:.2f}"
        elif isinstance(metric_value, int):
            formatted_value = f"{metric_value:,}"
        else:
            formatted_value = str(metric_value)
        
        summary_data.append({
            'Metric': display_name,
            'Value': formatted_value
        })
    
    return pd.DataFrame(summary_data)