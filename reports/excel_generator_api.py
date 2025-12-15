"""
Excel Report Generator for API Inspections with Photo Support
Building Inspection System V3 - Essential Community Management (ECM)

Features:
- Column G: Inspector Notes (50 chars wide, text wrap)
- Column H: Photo Thumbnails (150x150px)
- Downloads photos from SafetyCulture API
- Resizes to thumbnails using PIL
- Embeds using openpyxl
- Row height: 120 pixels for photo rows
"""

import os
import io
import requests
from datetime import datetime
from typing import Optional, List, Dict, Any
from io import BytesIO
from PIL import Image
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


class ExcelGeneratorAPI:
    """Generate Excel reports for API inspections with photo support"""
    
    def __init__(self, api_key: str):
        """
        Initialize the Excel generator
        
        Args:
            api_key: SafetyCulture API key for downloading photos
        """
        self.api_key = api_key
        self.photo_cache = {}  # Cache downloaded photos to avoid re-downloading
    
    def download_photo(self, photo_url: str) -> Optional[Image.Image]:
        """
        Download a photo from SafetyCulture API
        
        Args:
            photo_url: URL to the photo
            
        Returns:
            PIL Image object or None if download fails
        """
        # Check cache first
        if photo_url in self.photo_cache:
            return self.photo_cache[photo_url]
        
        try:
            headers = {'Authorization': f'Bearer {self.api_key}'}
            response = requests.get(photo_url, headers=headers, timeout=30)
            
            if response.status_code == 200:
                img = Image.open(BytesIO(response.content))
                self.photo_cache[photo_url] = img
                return img
            else:
                print(f"Failed to download photo: {photo_url} (Status: {response.status_code})")
                return None
                
        except Exception as e:
            print(f"Error downloading photo from {photo_url}: {str(e)}")
            return None
    
    def resize_to_thumbnail(self, img: Image.Image, size: tuple = (150, 150)) -> BytesIO:
        """
        Resize image to thumbnail maintaining aspect ratio
        
        Args:
            img: PIL Image object
            size: Target thumbnail size (width, height)
            
        Returns:
            BytesIO object containing the resized image
        """
        # Create a copy to avoid modifying the cached image
        img_copy = img.copy()
        
        # Calculate aspect ratio preserving resize
        img_copy.thumbnail(size, Image.Resampling.LANCZOS)
        
        # Save to BytesIO
        output = BytesIO()
        img_copy.save(output, format='PNG')
        output.seek(0)
        
        return output
    
    def generate_single_inspection_report(
        self,
        inspection_data: Dict[str, Any],
        defects: List[Dict[str, Any]],
        output_path: str
    ) -> bool:
        """
        Generate Excel report for a single inspection with photos
        
        Args:
            inspection_data: Dictionary containing inspection metadata
            defects: List of defect dictionaries with photo_url and inspector_notes
            output_path: Path where the Excel file should be saved
            
        Returns:
            True if successful, False otherwise
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Inspection Report"
            
            # Set column widths
            ws.column_dimensions['A'].width = 20  # Room
            ws.column_dimensions['B'].width = 25  # Component
            ws.column_dimensions['C'].width = 30  # Issue Description
            ws.column_dimensions['D'].width = 15  # Trade
            ws.column_dimensions['E'].width = 12  # Priority
            ws.column_dimensions['F'].width = 12  # Status
            ws.column_dimensions['G'].width = 50  # Inspector Notes
            ws.column_dimensions['H'].width = 20  # Photo
            
            # Header styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title Section
            ws.merge_cells('A1:H1')
            title_cell = ws['A1']
            title_cell.value = "BUILDING INSPECTION REPORT"
            title_cell.font = Font(size=16, bold=True, color="366092")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Inspection Details
            current_row = 3
            details = [
                ("Building:", inspection_data.get('building_name', 'N/A')),
                ("Inspection Date:", inspection_data.get('inspection_date', 'N/A')),
                ("Inspector:", inspection_data.get('inspector_name', 'N/A')),
                ("Total Defects:", str(len(defects))),
                ("With Photos:", str(sum(1 for d in defects if d.get('photo_url')))),
                ("With Notes:", str(sum(1 for d in defects if d.get('inspector_notes'))))
            ]
            
            for label, value in details:
                ws[f'A{current_row}'] = label
                ws[f'A{current_row}'].font = Font(bold=True)
                ws[f'B{current_row}'] = value
                ws.merge_cells(f'B{current_row}:C{current_row}')
                current_row += 1
            
            current_row += 1
            
            # Header Row
            headers = ['Room', 'Component', 'Issue Description', 'Trade', 'Priority', 'Status', 'Inspector Notes', 'Photo']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            header_row = current_row
            current_row += 1
            
            # Data Rows
            for defect in defects:
                row_start = current_row
                has_photo = bool(defect.get('photo_url'))
                
                # Set row height (120 pixels for photos, 30 for no photos)
                ws.row_dimensions[current_row].height = 120 if has_photo else 30
                
                # Cell alignment for data
                wrap_alignment = Alignment(vertical="top", wrap_text=True)
                center_alignment = Alignment(horizontal="center", vertical="center")
                
                # Room (A)
                cell = ws.cell(row=current_row, column=1)
                cell.value = defect.get('room', '')
                cell.alignment = wrap_alignment
                cell.border = thin_border
                
                # Component (B)
                cell = ws.cell(row=current_row, column=2)
                cell.value = defect.get('component', '')
                cell.alignment = wrap_alignment
                cell.border = thin_border
                
                # Issue Description (C)
                cell = ws.cell(row=current_row, column=3)
                cell.value = defect.get('issue_description', '')
                cell.alignment = wrap_alignment
                cell.border = thin_border
                
                # Trade (D)
                cell = ws.cell(row=current_row, column=4)
                cell.value = defect.get('trade', '')
                cell.alignment = center_alignment
                cell.border = thin_border
                
                # Priority (E)
                cell = ws.cell(row=current_row, column=5)
                priority = defect.get('priority', '')
                cell.value = priority
                cell.alignment = center_alignment
                cell.border = thin_border
                
                # Color coding for priority
                if priority == 'High':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif priority == 'Medium':
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif priority == 'Low':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                
                # Status (F)
                cell = ws.cell(row=current_row, column=6)
                cell.value = defect.get('status', 'Open')
                cell.alignment = center_alignment
                cell.border = thin_border
                
                # Inspector Notes (G)
                cell = ws.cell(row=current_row, column=7)
                cell.value = defect.get('inspector_notes', '')
                cell.alignment = wrap_alignment
                cell.border = thin_border
                
                # Photo (H)
                cell = ws.cell(row=current_row, column=8)
                cell.border = thin_border
                cell.alignment = center_alignment
                
                if has_photo:
                    photo_url = defect['photo_url']
                    img = self.download_photo(photo_url)
                    
                    if img:
                        # Resize to thumbnail
                        img_bytes = self.resize_to_thumbnail(img, size=(150, 150))
                        
                        # Create Excel image object
                        xl_img = XLImage(img_bytes)
                        
                        # Position the image in the cell
                        # Center it in the cell
                        cell_letter = get_column_letter(8)
                        xl_img.anchor = f'{cell_letter}{current_row}'
                        
                        # Add to worksheet
                        ws.add_image(xl_img)
                    else:
                        cell.value = "Photo unavailable"
                
                current_row += 1
            
            # Freeze header rows
            ws.freeze_panes = ws[f'A{header_row + 1}']
            
            # Auto-filter
            ws.auto_filter.ref = f'A{header_row}:H{current_row - 1}'
            
            # Save workbook
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"Error generating Excel report: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def generate_multi_inspection_report(
        self,
        inspections: List[Dict[str, Any]],
        output_path: str
    ) -> bool:
        """
        Generate Excel report for multiple inspections with photos
        
        Args:
            inspections: List of inspection dictionaries, each containing:
                - inspection_data: metadata
                - defects: list of defects with photos
            output_path: Path where the Excel file should be saved
            
        Returns:
            True if successful, False otherwise
        """
        try:
            wb = Workbook()
            
            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            # Header styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            
            # Summary title
            ws_summary['A1'] = "MULTI-INSPECTION REPORT SUMMARY"
            ws_summary['A1'].font = Font(size=16, bold=True, color="366092")
            ws_summary.merge_cells('A1:G1')
            ws_summary['A1'].alignment = Alignment(horizontal="center")
            
            # Summary headers
            summary_headers = ['Building', 'Inspection Date', 'Inspector', 'Total Defects', 'With Photos', 'With Notes', 'Status']
            for col_num, header in enumerate(summary_headers, 1):
                cell = ws_summary.cell(row=3, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Summary data
            row = 4
            total_defects = 0
            total_photos = 0
            total_notes = 0
            
            for inspection in inspections:
                data = inspection['inspection_data']
                defects = inspection['defects']
                
                defect_count = len(defects)
                photo_count = sum(1 for d in defects if d.get('photo_url'))
                note_count = sum(1 for d in defects if d.get('inspector_notes'))
                
                total_defects += defect_count
                total_photos += photo_count
                total_notes += note_count
                
                ws_summary.cell(row=row, column=1, value=data.get('building_name', 'N/A'))
                ws_summary.cell(row=row, column=2, value=data.get('inspection_date', 'N/A'))
                ws_summary.cell(row=row, column=3, value=data.get('inspector_name', 'N/A'))
                ws_summary.cell(row=row, column=4, value=defect_count)
                ws_summary.cell(row=row, column=5, value=photo_count)
                ws_summary.cell(row=row, column=6, value=note_count)
                ws_summary.cell(row=row, column=7, value="Complete")
                row += 1
            
            # Totals row
            row += 1
            ws_summary.cell(row=row, column=1, value="TOTALS").font = Font(bold=True)
            ws_summary.cell(row=row, column=4, value=total_defects).font = Font(bold=True)
            ws_summary.cell(row=row, column=5, value=total_photos).font = Font(bold=True)
            ws_summary.cell(row=row, column=6, value=total_notes).font = Font(bold=True)
            
            # Set column widths
            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 15
            ws_summary.column_dimensions['C'].width = 20
            ws_summary.column_dimensions['D'].width = 15
            ws_summary.column_dimensions['E'].width = 15
            ws_summary.column_dimensions['F'].width = 15
            ws_summary.column_dimensions['G'].width = 12
            
            # Create individual sheets for each inspection
            for idx, inspection in enumerate(inspections, 1):
                sheet_name = f"Inspection_{idx}"
                ws = wb.create_sheet(title=sheet_name)
                
                data = inspection['inspection_data']
                defects = inspection['defects']
                
                # Use same format as single inspection
                self._add_inspection_to_sheet(ws, data, defects)
            
            # Save workbook
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"Error generating multi-inspection Excel report: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def _add_inspection_to_sheet(
        self,
        ws,
        inspection_data: Dict[str, Any],
        defects: List[Dict[str, Any]]
    ):
        """Helper method to add inspection data to a worksheet"""
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 50
        ws.column_dimensions['H'].width = 20
        
        # Header styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = f"Inspection: {inspection_data.get('building_name', 'N/A')}"
        ws['A1'].font = Font(size=14, bold=True, color="366092")
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Details
        current_row = 3
        ws[f'A{current_row}'] = "Inspection Date:"
        ws[f'B{current_row}'] = inspection_data.get('inspection_date', 'N/A')
        current_row += 1
        ws[f'A{current_row}'] = "Inspector:"
        ws[f'B{current_row}'] = inspection_data.get('inspector_name', 'N/A')
        current_row += 2
        
        # Headers
        headers = ['Room', 'Component', 'Issue Description', 'Trade', 'Priority', 'Status', 'Inspector Notes', 'Photo']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        current_row += 1
        
        # Data rows (same logic as single inspection)
        for defect in defects:
            has_photo = bool(defect.get('photo_url'))
            ws.row_dimensions[current_row].height = 120 if has_photo else 30
            
            wrap_alignment = Alignment(vertical="top", wrap_text=True)
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # Add all cell data
            ws.cell(row=current_row, column=1, value=defect.get('room', '')).alignment = wrap_alignment
            ws.cell(row=current_row, column=2, value=defect.get('component', '')).alignment = wrap_alignment
            ws.cell(row=current_row, column=3, value=defect.get('issue_description', '')).alignment = wrap_alignment
            ws.cell(row=current_row, column=4, value=defect.get('trade', '')).alignment = center_alignment
            
            priority_cell = ws.cell(row=current_row, column=5, value=defect.get('priority', ''))
            priority_cell.alignment = center_alignment
            priority = defect.get('priority', '')
            if priority == 'High':
                priority_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif priority == 'Medium':
                priority_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif priority == 'Low':
                priority_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            ws.cell(row=current_row, column=6, value=defect.get('status', 'Open')).alignment = center_alignment
            ws.cell(row=current_row, column=7, value=defect.get('inspector_notes', '')).alignment = wrap_alignment
            
            # Apply borders to all cells
            for col in range(1, 9):
                ws.cell(row=current_row, column=col).border = thin_border
            
            # Photo
            if has_photo:
                photo_url = defect['photo_url']
                img = self.download_photo(photo_url)
                
                if img:
                    img_bytes = self.resize_to_thumbnail(img, size=(150, 150))
                    xl_img = XLImage(img_bytes)
                    cell_letter = get_column_letter(8)
                    xl_img.anchor = f'{cell_letter}{current_row}'
                    ws.add_image(xl_img)
                else:
                    ws.cell(row=current_row, column=8, value="Photo unavailable")
            
            current_row += 1


def create_excel_report_from_database(
    inspection_ids: List[int],
    db_connection,
    api_key: str,
    output_path: str,
    report_type: str = "single"
) -> bool:
    """
    Main function to create Excel report from database data
    
    Args:
        inspection_ids: List of inspection IDs to include
        db_connection: Database connection object
        api_key: SafetyCulture API key
        output_path: Where to save the Excel file
        report_type: "single" or "multi" inspection report
        
    Returns:
        True if successful, False otherwise
    """
    try:
        generator = ExcelGeneratorAPI(api_key)
        
        if report_type == "single" and len(inspection_ids) == 1:
            # Query single inspection
            inspection_data, defects = _query_inspection_data(db_connection, inspection_ids[0])
            return generator.generate_single_inspection_report(inspection_data, defects, output_path)
        
        elif report_type == "multi":
            # Query multiple inspections
            inspections = []
            for inspection_id in inspection_ids:
                inspection_data, defects = _query_inspection_data(db_connection, inspection_id)
                inspections.append({
                    'inspection_data': inspection_data,
                    'defects': defects
                })
            return generator.generate_multi_inspection_report(inspections, output_path)
        
        else:
            print(f"Invalid report type or inspection count: {report_type}, {len(inspection_ids)} inspections")
            return False
            
    except Exception as e:
        print(f"Error creating Excel report: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def _query_inspection_data(db_connection, inspection_id: int) -> tuple:
    """
    Query inspection data and defects from database
    
    Args:
        db_connection: Database connection
        inspection_id: ID of the inspection
        
    Returns:
        Tuple of (inspection_data dict, defects list)
    """
    cursor = db_connection.cursor()
    
    # Query inspection metadata
    cursor.execute("""
        SELECT i.id, i.inspection_date, i.inspector_name, i.total_defects,
               b.name as building_name
        FROM inspector_inspections i
        JOIN inspector_buildings b ON i.building_id = b.id
        WHERE i.id = %s
    """, (inspection_id,))
    
    row = cursor.fetchone()
    if not row:
        raise ValueError(f"Inspection {inspection_id} not found")
    
    inspection_data = {
        'id': row[0],
        'inspection_date': row[1].strftime('%Y-%m-%d') if row[1] else 'N/A',
        'inspector_name': row[2] or 'N/A',
        'total_defects': row[3],
        'building_name': row[4]
    }
    
    # Query defects
    cursor.execute("""
        SELECT room, component, issue_description, trade, priority, status,
               inspector_notes, photo_url, photo_media_id
        FROM inspector_inspection_items
        WHERE inspection_id = %s
        ORDER BY room, component
    """, (inspection_id,))
    
    defects = []
    for row in cursor.fetchall():
        defects.append({
            'room': row[0] or '',
            'component': row[1] or '',
            'issue_description': row[2] or '',
            'trade': row[3] or '',
            'priority': row[4] or '',
            'status': row[5] or 'Open',
            'inspector_notes': row[6] or '',
            'photo_url': row[7] or '',
            'photo_media_id': row[8] or ''
        })
    
    cursor.close()
    return inspection_data, defects