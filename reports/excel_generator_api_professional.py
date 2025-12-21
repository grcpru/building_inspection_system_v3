"""
Professional Excel Report Generator for API Inspections with Photos
Building Inspection System V3 - Essential Community Management (ECM)

Combines:
- Professional template from excel_report_generator.py (dashboards, metrics, analysis)
- Photo support from excel_generator_api.py (SafetyCulture API integration)
- Database integration for PostgreSQL/Supabase

Features:
- Executive Dashboard with Quality Score
- Settlement Readiness Analysis
- Multiple summary sheets (Trade, Room, Component, Unit)
- Component Details with units affected
- Photos embedded as thumbnails in data sheets
- Professional formatting and color coding
- Workflow Tracker with status monitoring
"""

import os
import io
import requests
from datetime import datetime
from typing import Optional, List, Dict, Any
from io import BytesIO
from PIL import Image
import pandas as pd
import pytz
import xlsxwriter
import logging

# Set up logging
logger = logging.getLogger(__name__)


class ProfessionalExcelGeneratorAPI:
    """Generate professional Excel reports for API inspections with photos"""
    
    def __init__(self, api_key: str):
        """
        Initialize the professional Excel generator
        
        Args:
            api_key: SafetyCulture API key for downloading photos
        """
        self.api_key = api_key
        self.photo_cache = {}  # Cache downloaded photos
    
    def download_photo(self, photo_url: str) -> Optional[Image.Image]:
        """
        Download a photo from SafetyCulture API
        
        Args:
            photo_url: URL to the photo
            
        Returns:
            PIL Image object or None if download fails
        """
        # Check cache first
        if photo_url in self.photo_cache:
            return self.photo_cache[photo_url]
        
        try:
            headers = {'Authorization': f'Bearer {self.api_key}'}
            response = requests.get(photo_url, headers=headers, timeout=30)
            
            if response.status_code == 200:
                img = Image.open(BytesIO(response.content))
                self.photo_cache[photo_url] = img
                return img
            else:
                logger.warning(f"Failed to download photo: {photo_url} (Status: {response.status_code})")
                return None
                
        except Exception as e:
            logger.error(f"Error downloading photo from {photo_url}: {str(e)}")
            return None
    
    def resize_to_thumbnail(self, img: Image.Image, size: tuple = (150, 150)) -> BytesIO:
        """
        Resize image to thumbnail maintaining aspect ratio
        
        Args:
            img: PIL Image object
            size: Target thumbnail size (width, height)
            
        Returns:
            BytesIO object containing the resized image
        """
        # Create a copy to avoid modifying the cached image
        img_copy = img.copy()
        
        # Calculate aspect ratio preserving resize
        img_copy.thumbnail(size, Image.Resampling.LANCZOS)
        
        # Save to BytesIO
        output = BytesIO()
        img_copy.save(output, format='PNG')
        output.seek(0)
        
        return output
    
    def transform_api_data(self, inspection_data: Dict[str, Any], defects: List[Dict[str, Any]]) -> tuple:
        """
        Transform API data to DataFrame format expected by professional template
        
        Args:
            inspection_data: Inspection metadata dict
            defects: List of defect dicts from database
            
        Returns:
            tuple: (processed_data DataFrame, metrics dict)
        """
        # Convert defects to DataFrame
        if len(defects) == 0:
            processed_data = pd.DataFrame(columns=[
                'Unit', 'UnitType', 'Room', 'Component', 'Trade', 'StatusClass',
                'Urgency', 'InspectionDate', 'InspectorNotes', 'IssueDescription',
                'photo_url', 'photo_media_id'
            ])
        else:
            processed_data = pd.DataFrame(defects)
            
            # Rename columns to match template
            column_mapping = {
                'room': 'Room',
                'component': 'Component',
                'trade': 'Trade',
                'priority': 'Urgency',
                'status': 'StatusClass',
                'inspector_notes': 'InspectorNotes',
                'unit': 'Unit',  # Map lowercase unit to uppercase Unit
                'unit_type': 'UnitType'  # Map unit_type as well
            }
            
            for old_col, new_col in column_mapping.items():
                if old_col in processed_data.columns:
                    processed_data[new_col] = processed_data[old_col]
            
            # Only set default Unit/UnitType if not already set
            if 'Unit' not in processed_data.columns:
                # Extract from inspection_data as fallback
                unit_name = inspection_data.get('unit', 'Unit')
                processed_data['Unit'] = unit_name
            
            if 'UnitType' not in processed_data.columns:
                unit_type = inspection_data.get('unit_type', 'Apartment')
                processed_data['UnitType'] = unit_type
            
            # CRITICAL FIX: Use 'notes' field for IssueDescription (this has the actual defect description!)
            if 'description' in processed_data.columns:
                processed_data['IssueDescription'] = processed_data['description']
            elif 'notes' in processed_data.columns:
                processed_data['IssueDescription'] = processed_data['notes']
            else:
                processed_data['IssueDescription'] = 'No description'
            
            # Preserve date columns for workflow tracker (keep original names)
            # These are already in the DataFrame from the query
            # inspection_date, created_at, planned_completion, owner_signoff_timestamp
            # unit, building_name, inspection_id - also preserved
            
            # Add inspection_id if not present (use from inspection_data)
            if 'inspection_id' not in processed_data.columns:
                processed_data['inspection_id'] = inspection_data.get('id', '')
            
            # Add InspectionDate column if not present
            if 'InspectionDate' not in processed_data.columns:
                if 'inspection_date' in processed_data.columns:
                    processed_data['InspectionDate'] = processed_data['inspection_date']
                else:
                    insp_date = inspection_data.get('inspection_date', datetime.now())
                    if isinstance(insp_date, str):
                        insp_date = pd.to_datetime(insp_date)
                    processed_data['InspectionDate'] = insp_date
            
            # Ensure StatusClass is set correctly (normalize to "Not OK")
            if 'StatusClass' not in processed_data.columns:
                if 'status' in processed_data.columns:
                    processed_data['StatusClass'] = processed_data['status']
                else:
                    processed_data['StatusClass'] = 'Not OK'
            
            # Normalize StatusClass values to "Not OK" (handle case variations)
            if 'StatusClass' in processed_data.columns:
                processed_data['StatusClass'] = processed_data['StatusClass'].str.strip()
                # Ensure all defects have "Not OK" status (they came from status_class = 'not ok' filter)
                processed_data.loc[processed_data['StatusClass'].notna(), 'StatusClass'] = 'Not OK'
        
        # Calculate metrics
        metrics = self.calculate_metrics(inspection_data, processed_data)
        
        return processed_data, metrics
    
    def calculate_metrics(self, inspection_data: Dict[str, Any], processed_data: pd.DataFrame) -> Dict[str, Any]:
        """
        Calculate all metrics needed for professional dashboard
        
        Args:
            inspection_data: Inspection metadata
            processed_data: DataFrame with defect data
            
        Returns:
            Dict with all calculated metrics
        """
        # Basic counts
        total_defects = len(processed_data[processed_data['StatusClass'] == 'Not OK'])
        total_inspections = inspection_data.get('total_items', len(processed_data))
        
        # Defect rate
        defect_rate = (total_defects / total_inspections * 100) if total_inspections > 0 else 0
        
        # Settlement readiness (per unit analysis)
        if 'Unit' in processed_data.columns and len(processed_data) > 0:
            # Count units with defects
            unit_defect_counts = processed_data[processed_data['StatusClass'] == 'Not OK'].groupby('Unit').size()
            
            # Calculate readiness categories
            ready_units = len(unit_defect_counts[unit_defect_counts <= 2])
            minor_work_units = len(unit_defect_counts[(unit_defect_counts >= 3) & (unit_defect_counts <= 7)])
            major_work_units = len(unit_defect_counts[(unit_defect_counts >= 8) & (unit_defect_counts <= 15)])
            extensive_work_units = len(unit_defect_counts[unit_defect_counts > 15])
            
            # Total units = unique units in the data (for multi-inspection, this counts all inspected units)
            total_units = processed_data['Unit'].nunique()
        else:
            # Single unit inspection
            ready_units = 1 if total_defects <= 2 else 0
            minor_work_units = 1 if 3 <= total_defects <= 7 else 0
            major_work_units = 1 if 8 <= total_defects <= 15 else 0
            extensive_work_units = 1 if total_defects > 15 else 0
            total_units = 1
        
        # Calculate percentages
        ready_pct = (ready_units / total_units * 100) if total_units > 0 else 0
        minor_pct = (minor_work_units / total_units * 100) if total_units > 0 else 0
        major_pct = (major_work_units / total_units * 100) if total_units > 0 else 0
        extensive_pct = (extensive_work_units / total_units * 100) if total_units > 0 else 0
        
        # Trade summary
        if len(processed_data[processed_data['StatusClass'] == 'Not OK']) > 0:
            summary_trade = processed_data[processed_data['StatusClass'] == 'Not OK'].groupby('Trade').size().reset_index(name='DefectCount')
            summary_trade = summary_trade.sort_values('DefectCount', ascending=False)
        else:
            summary_trade = pd.DataFrame(columns=['Trade', 'DefectCount'])
        
        # Room summary
        if 'Room' in processed_data.columns and len(processed_data[processed_data['StatusClass'] == 'Not OK']) > 0:
            summary_room = processed_data[processed_data['StatusClass'] == 'Not OK'].groupby('Room').size().reset_index(name='DefectCount')
            summary_room = summary_room.sort_values('DefectCount', ascending=False)
        else:
            summary_room = pd.DataFrame(columns=['Room', 'DefectCount'])
        
        # Component summary
        if 'Component' in processed_data.columns and len(processed_data[processed_data['StatusClass'] == 'Not OK']) > 0:
            summary_component = processed_data[processed_data['StatusClass'] == 'Not OK'].groupby('Component').size().reset_index(name='DefectCount')
            summary_component = summary_component.sort_values('DefectCount', ascending=False)
        else:
            summary_component = pd.DataFrame(columns=['Component', 'DefectCount'])
        
        # Unit summary
        if 'Unit' in processed_data.columns and len(processed_data[processed_data['StatusClass'] == 'Not OK']) > 0:
            summary_unit = processed_data[processed_data['StatusClass'] == 'Not OK'].groupby('Unit').size().reset_index(name='DefectCount')
            summary_unit = summary_unit.sort_values('DefectCount', ascending=False)
        else:
            summary_unit = pd.DataFrame(columns=['Unit', 'DefectCount'])
        
        # Photo and notes counts
        photo_count = processed_data['photo_url'].notna().sum() if 'photo_url' in processed_data.columns else 0
        notes_count = processed_data['InspectorNotes'].notna().sum() if 'InspectorNotes' in processed_data.columns else 0
        
        # Extract unique unit types from actual data
        if 'UnitType' in processed_data.columns and len(processed_data) > 0:
            unique_unit_types = processed_data['UnitType'].dropna().unique()
            if len(unique_unit_types) > 0:
                unit_types_str = ', '.join(sorted(set(str(ut) for ut in unique_unit_types)))
            else:
                unit_types_str = inspection_data.get('unit_type', 'Apartment')
        else:
            unit_types_str = inspection_data.get('unit_type', 'Apartment')
        
        # Extract address from API metadata if available
        # SafetyCulture stores address in fields like "Title Page_Site conducted_Area", "Title Page_Site conducted_Region"
        address = inspection_data.get('address', '')
        
        # If address is empty, try multiple sources
        if not address or address == '':
            # Try 1: Check inspection_data metadata
            metadata = inspection_data.get('metadata', {})
            if metadata:
                area = metadata.get('Title Page_Site conducted_Area') or metadata.get('site_area') or ''
                region = metadata.get('Title Page_Site conducted_Region') or metadata.get('site_region') or ''
                
                if area and region:
                    address = f"{area}, {region}"
                elif area:
                    address = area
                elif region:
                    address = region
            
            # Try 2: Check if defects have address fields
            if (not address or address == '') and len(processed_data) > 0:
                # Check for address-like columns in the data
                if 'site_area' in processed_data.columns:
                    area = processed_data['site_area'].dropna().iloc[0] if len(processed_data['site_area'].dropna()) > 0 else ''
                    if area:
                        address = str(area)
                
                if 'site_region' in processed_data.columns and address:
                    region = processed_data['site_region'].dropna().iloc[0] if len(processed_data['site_region'].dropna()) > 0 else ''
                    if region:
                        address = f"{address}, {region}"
            
            # Fallback
            if not address or address == '':
                address = 'Address not specified'
        
        # Build metrics dictionary
        metrics = {
            'building_name': inspection_data.get('building_name', 'Building'),
            'address': address,  # Use extracted/calculated address
            'inspection_date': inspection_data.get('inspection_date', datetime.now().strftime('%Y-%m-%d')),
            'unit_types_str': unit_types_str,  # Use calculated unit types from data
            'total_units': total_units,
            'total_inspections': total_inspections,
            'total_defects': total_defects,
            'defect_rate': defect_rate,
            'avg_defects_per_unit': total_defects / total_units if total_units > 0 else 0,
            'ready_units': ready_units,
            'minor_work_units': minor_work_units,
            'major_work_units': major_work_units,
            'extensive_work_units': extensive_work_units,
            'ready_pct': ready_pct,
            'minor_pct': minor_pct,
            'major_pct': major_pct,
            'extensive_pct': extensive_pct,
            'summary_trade': summary_trade,
            'summary_room': summary_room,
            'summary_component': summary_component,
            'summary_unit': summary_unit,
            'photo_count': photo_count,
            'notes_count': notes_count,
            'is_multi_day_inspection': False
        }
        
        return metrics
    
    def generate_professional_report(
        self,
        inspection_data: Dict[str, Any],
        defects: List[Dict[str, Any]],
        output_path: str
    ) -> bool:
        """
        Generate professional Excel report with photos
        
        Args:
            inspection_data: Inspection metadata dict
            defects: List of defect dicts with photo URLs
            output_path: Path to save Excel file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            logger.info("Starting professional Excel report generation")
            
            # Transform data
            processed_data, metrics = self.transform_api_data(inspection_data, defects)
            
            logger.info(f"Transformed data: {len(processed_data)} rows, {len(metrics['summary_trade'])} trades")
            
            # Generate Excel using two-pass process
            success = self._generate_excel_with_photos(processed_data, metrics, output_path)
            
            if success:
                logger.info(f"Professional Excel report saved to: {output_path}")
                return True
            else:
                logger.error("Failed to generate professional Excel report")
                return False
            
        except Exception as e:
            logger.error(f"Error generating professional Excel report: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def _generate_excel_with_photos(self, final_df: pd.DataFrame, metrics: dict, temp_path: str) -> bool:
        """
        Two-pass Excel generation with professional template + photos
        
        Pass 1: xlsxwriter creates professional dashboard and formatting
        Pass 2: openpyxl adds photo thumbnails
        
        Args:
            final_df: DataFrame with defect data
            metrics: Calculated metrics dict
            temp_path: Temporary file path for intermediate Excel
            
        Returns:
            True if successful
        """
        logger.info("Starting two-pass Excel generation (xlsxwriter + openpyxl)")
        
        try:
            # ===== PASS 1: xlsxwriter - Professional Template =====
            logger.info("Pass 1: Creating professional template with xlsxwriter")
            
            # Create workbook with xlsxwriter
            workbook = xlsxwriter.Workbook(temp_path, {
                'nan_inf_to_errors': True,
                'remove_timezone': True
            })
            
            # Define all formats
            formats = self._create_formats(workbook)
            
            # 1. Executive Dashboard
            self._create_executive_dashboard(workbook, metrics, formats)
            
            # 2. Defects Only (with photos - renamed from All Defects)
            defects_sheet_idx = self._create_data_sheet_with_photos(workbook, final_df, "⚠️ Defects Only", formats)
            
            # 3. All Inspections (all items, no photos)
            self._create_all_inspections_sheet(workbook, final_df, metrics, formats)
            
            # 4. Component Details (analysis by component)
            self._create_component_details_sheet(workbook, final_df, formats)
            
            # Debug: Log what columns exist and summary DataFrame info
            logger.info(f"DataFrame columns after transform: {list(final_df.columns)}")
            logger.info(f"Trade summary shape: {metrics.get('summary_trade', pd.DataFrame()).shape}")
            logger.info(f"Room summary shape: {metrics.get('summary_room', pd.DataFrame()).shape}")
            logger.info(f"Component summary shape: {metrics.get('summary_component', pd.DataFrame()).shape}")
            logger.info(f"Unit summary shape: {metrics.get('summary_unit', pd.DataFrame()).shape}")
            
            # 5. Trade Summary
            if len(metrics.get('summary_trade', pd.DataFrame())) > 0:
                logger.info("Creating Trade Summary sheet")
                self._create_summary_sheet(workbook, metrics['summary_trade'], "🔧 Trade Summary", formats)
            else:
                logger.warning("Skipping Trade Summary - empty DataFrame")
            
            # 6. Room Summary
            if len(metrics.get('summary_room', pd.DataFrame())) > 0:
                logger.info("Creating Room Summary sheet")
                self._create_summary_sheet(workbook, metrics['summary_room'], "🚪 Room Summary", formats)
            else:
                logger.warning("Skipping Room Summary - empty DataFrame")
            
            # 7. Component Summary
            if len(metrics.get('summary_component', pd.DataFrame())) > 0:
                logger.info("Creating Component Summary sheet")
                self._create_summary_sheet(workbook, metrics['summary_component'], "🔩 Component Summary", formats)
            else:
                logger.warning("Skipping Component Summary - empty DataFrame")
            
            # 8. Unit Summary
            if len(metrics.get('summary_unit', pd.DataFrame())) > 0:
                logger.info("Creating Unit Summary sheet")
                self._create_summary_sheet(workbook, metrics['summary_unit'], "🏠 Unit Summary", formats)
            else:
                logger.warning("Skipping Unit Summary - empty DataFrame")
            
            # 9. Metadata
            self._create_metadata_sheet(workbook, metrics, formats)
            
            # 10. Workflow Tracker
            self._create_workflow_tracker_sheet(workbook, final_df, metrics, formats)
            
            # Close xlsxwriter workbook
            workbook.close()
            logger.info("Pass 1 complete: Professional template created")
            
            # ===== PASS 2: openpyxl - Add Photos =====
            logger.info("Pass 2: Adding photos with openpyxl")
            self._add_photos_with_openpyxl(temp_path, final_df)
            logger.info("Pass 2 complete: Photos embedded")
            
            return True
            
        except Exception as e:
            logger.error(f"Error in two-pass Excel generation: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def _create_formats(self, workbook):
        """Create all formats needed for professional template"""
        return {
            'title': workbook.add_format({
                'bold': True,
                'font_size': 18,
                'bg_color': '#4CAF50',
                'font_color': 'white',
                'align': 'center',
                'valign': 'vcenter',
                'border': 2
            }),
            'header': workbook.add_format({
                'bold': True,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': '#1F4E78',
                'font_color': 'white',
                'border': 1
            }),
            'cell': workbook.add_format({
                'align': 'left',
                'valign': 'vcenter',
                'border': 1,
                'font_size': 10
            }),
            'cell_alt': workbook.add_format({
                'align': 'left',
                'valign': 'vcenter',
                'border': 1,
                'font_size': 10,
                'bg_color': '#F7F9FC'
            }),
            'notes': workbook.add_format({
                'align': 'left',
                'valign': 'top',
                'border': 1,
                'text_wrap': True,
                'font_size': 10
            }),
            'notes_alt': workbook.add_format({
                'align': 'left',
                'valign': 'top',
                'border': 1,
                'text_wrap': True,
                'font_size': 10,
                'bg_color': '#F7F9FC'
            }),
            'label': workbook.add_format({
                'bold': True,
                'font_size': 11,
                'bg_color': '#F5F5F5',
                'border': 1
            }),
            'data': workbook.add_format({
                'font_size': 11,
                'border': 1,
                'align': 'right'
            }),
            'ready': workbook.add_format({
                'font_size': 11,
                'border': 1,
                'align': 'right',
                'bg_color': '#C8E6C9',
                'font_color': '#2E7D32'
            }),
            'minor': workbook.add_format({
                'font_size': 11,
                'border': 1,
                'align': 'right',
                'bg_color': '#FFF3C4',
                'font_color': '#F57F17'
            }),
            'major': workbook.add_format({
                'font_size': 11,
                'border': 1,
                'align': 'right',
                'bg_color': '#FFCDD2',
                'font_color': '#C62828'
            }),
            'extensive': workbook.add_format({
                'font_size': 11,
                'border': 1,
                'align': 'right',
                'bg_color': '#F8BBD9',
                'font_color': '#AD1457'
            }),
            'quality_score': workbook.add_format({
                'font_size': 11,
                'border': 1,
                'align': 'right',
                'bg_color': '#C8E6C9',
                'font_color': '#2E7D32',
                'bold': True
            })
        }
    
    def _create_executive_dashboard(self, workbook, metrics, formats):
        """Create executive dashboard sheet matching template format"""
        ws = workbook.add_worksheet("📊 Executive Dashboard")
        ws.set_column('A:A', 35)
        ws.set_column('B:B', 45)
        
        row = 0
        
        # Title with building name
        ws.merge_range(f'A{row+1}:B{row+1}', 
                       f'🏢 {metrics["building_name"].upper()} - INSPECTION REPORT',
                       formats['title'])
        ws.set_row(row, 30)
        row += 2
        
        # ===== BUILDING INFORMATION SECTION =====
        info_header = workbook.add_format({
            'bold': True,
            'font_size': 12,
            'bg_color': '#4472C4',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        ws.merge_range(f'A{row+1}:B{row+1}', '🏢 BUILDING INFORMATION', info_header)
        row += 1
        
        building_data = [
            ('Building Name', metrics['building_name']),
            ('Address', metrics['address']),
            ('Inspection Date', metrics['inspection_date']),
            ('Total Units Inspected', f"{metrics['total_units']:,}"),
            ('Unit Types', metrics['unit_types_str'])
        ]
        
        for label, value in building_data:
            ws.write(row, 0, label, formats['label'])
            ws.write(row, 1, value, formats['data'])
            row += 1
        
        row += 1
        
        # ===== INSPECTION SUMMARY SECTION =====
        summary_header = workbook.add_format({
            'bold': True,
            'font_size': 12,
            'bg_color': '#F4B084',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        ws.merge_range(f'A{row+1}:B{row+1}', '📊 INSPECTION SUMMARY', summary_header)
        row += 1
        
        quality_score = max(0, 100 - metrics.get('defect_rate', 0))
        
        inspection_data = [
            ('Total Inspection Points', f"{metrics['total_inspections']:,}", formats['data']),
            ('Total Defects Found', f"{metrics['total_defects']:,}", formats['data']),
            ('Overall Defect Rate', f"{metrics['defect_rate']:.2f}%", formats['data']),
            ('Average Defects per Unit', f"{metrics['avg_defects_per_unit']:.1f}", formats['data']),
            ('Development Quality Score', f"{quality_score:.1f}/100", formats['quality_score'])
        ]
        
        for label, value, fmt in inspection_data:
            ws.write(row, 0, label, formats['label'])
            ws.write(row, 1, value, fmt)
            row += 1
        
        row += 1
        
        # ===== SETTLEMENT READINESS ANALYSIS SECTION =====
        readiness_header = workbook.add_format({
            'bold': True,
            'font_size': 12,
            'bg_color': '#70AD47',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        ws.merge_range(f'A{row+1}:B{row+1}', '🏠 SETTLEMENT READINESS ANALYSIS', readiness_header)
        row += 1
        
        readiness_data = [
            ('☑ Minor Work Required (0-2 defects)',
             f"{metrics['ready_units']} units ({metrics['ready_pct']:.1f}%)", formats['ready']),
            ('⚠ Intermediate Remediation Required (3-7 defects)',
             f"{metrics['minor_work_units']} units ({metrics['minor_pct']:.1f}%)", formats['minor']),
            ('🔨 Major Work Required (8-15 defects)',
             f"{metrics['major_work_units']} units ({metrics['major_pct']:.1f}%)", formats['major']),
            ('🚧 Extensive Work Required (15+ defects)',
             f"{metrics['extensive_work_units']} units ({metrics['extensive_pct']:.1f}%)", formats['extensive'])
        ]
        
        for label, value, fmt in readiness_data:
            ws.write(row, 0, label, formats['label'])
            ws.write(row, 1, value, fmt)
            row += 1
        
        row += 1
        
        # ===== QUALITY SCORE ANALYSIS =====
        quality_header = workbook.add_format({
            'bold': True,
            'font_size': 12,
            'bg_color': '#FFC000',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        ws.merge_range(f'A{row+1}:B{row+1}', '⚙ QUALITY SCORE ANALYSIS', quality_header)
        row += 1
        
        # Component pass rate calculation
        component_pass_rate = 100 - metrics.get('defect_rate', 0)
        
        # Quality grade based on score
        if quality_score >= 98:
            quality_grade = 'Excellent (A+)'
        elif quality_score >= 95:
            quality_grade = 'Very Good (A)'
        elif quality_score >= 90:
            quality_grade = 'Good (B+)'
        elif quality_score >= 85:
            quality_grade = 'Satisfactory (B)'
        else:
            quality_grade = 'Needs Improvement (C)'
        
        # Industry benchmark
        if quality_score >= 98:
            benchmark = 'Above Industry Standard'
        elif quality_score >= 95:
            benchmark = 'Meets Industry Standard'
        else:
            benchmark = 'Below Industry Standard'
        
        # Recommended action
        if metrics['total_defects'] == 0:
            action = 'Ready for settlement'
        elif metrics['avg_defects_per_unit'] <= 3:
            action = 'Minor remediation required'
        elif metrics['avg_defects_per_unit'] <= 8:
            action = 'Moderate remediation required'
        else:
            action = 'Significant remediation required'
        
        quality_data = [
            ('Component Pass Rate', f"{component_pass_rate:.1f}%", formats['quality_score']),
            ('Quality Grade', quality_grade, formats['data']),
            ('Industry Benchmark', benchmark, formats['data']),
            ('Recommended Action', action, formats['data'])
        ]
        
        for label, value, fmt in quality_data:
            ws.write(row, 0, label, formats['label'])
            ws.write(row, 1, value, fmt)
            row += 1
        
        row += 1
        
        # ===== TOP PROBLEM TRADES =====
        trades_header = workbook.add_format({
            'bold': True,
            'font_size': 12,
            'bg_color': '#C55A11',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        ws.merge_range(f'A{row+1}:B{row+1}', '⚠ TOP PROBLEM TRADES', trades_header)
        row += 1
        
        if len(metrics.get('summary_trade', pd.DataFrame())) > 0:
            top_trades = metrics['summary_trade'].head(10)
            for idx, (_, trade_row) in enumerate(top_trades.iterrows(), 1):
                trade_label = f"{idx}. {trade_row['Trade']}"
                defect_count = f"{trade_row['DefectCount']} defects"
                ws.write(row, 0, trade_label, formats['label'])
                ws.write(row, 1, defect_count, formats['data'])
                row += 1
        
        # Footer
        row += 2
        melbourne_tz = pytz.timezone('Australia/Melbourne')
        melbourne_time = datetime.now(melbourne_tz)
        report_time = melbourne_time.strftime('%d/%m/%Y at %I:%M %p AEDT')
        
        ws.merge_range(f'A{row+1}:B{row+1}',
                       f'Report generated on {report_time} | Professional Inspection Report with Photos',
                       workbook.add_format({'font_size': 9, 'italic': True, 'align': 'center'}))
    
    def _create_data_sheet_with_photos(self, workbook, data_df, sheet_name, formats):
        """
        Create defects-only sheet with photos (Pass 1)
        Format: Inspection Date | Building | Unit | Room | Component | Trade | Priority | Status | Inspector Notes | Photo
        Photos will be added in Pass 2 with openpyxl
        
        Returns:
            Sheet name for photo embedding in Pass 2
        """
        ws = workbook.add_worksheet(sheet_name)
        
        # Column widths - Inspection Date FIRST, NO Issue Description
        ws.set_column('A:A', 15)  # Inspection Date
        ws.set_column('B:B', 15)  # Building
        ws.set_column('C:C', 12)  # Unit
        ws.set_column('D:D', 20)  # Room
        ws.set_column('E:E', 25)  # Component
        ws.set_column('F:F', 15)  # Trade
        ws.set_column('G:G', 12)  # Priority
        ws.set_column('H:H', 12)  # Status
        ws.set_column('I:I', 50)  # Inspector Notes
        ws.set_column('J:J', 20)  # Photo (will be populated in Pass 2)
        
        # Headers - Inspection Date FIRST, no Issue Description
        headers = ['Inspection Date', 'Building', 'Unit', 'Room', 'Component', 'Trade', 'Priority', 'Status', 'Inspector Notes', 'Photo']
        for col_idx, header in enumerate(headers):
            ws.write(0, col_idx, header, formats['header'])
        
        # Date format
        date_fmt = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'num_format': 'dd/mm/yyyy'
        })
        date_fmt_alt = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'num_format': 'dd/mm/yyyy',
            'bg_color': '#F7F9FC'
        })
        
        # Data rows
        for row_idx, (_, row) in enumerate(data_df.iterrows(), start=1):
            is_alt = (row_idx % 2 == 0)
            base_fmt = formats['cell_alt'] if is_alt else formats['cell']
            notes_fmt = formats['notes_alt'] if is_alt else formats['notes']
            date_row_fmt = date_fmt_alt if is_alt else date_fmt
            
            # Get unit for building determination
            unit = str(row.get('unit') or row.get('Unit', ''))
            
            # Building - Simplified based on unit prefix
            if unit and len(unit) > 0:
                first_char = unit[0].upper()
                if first_char == 'G':
                    building_display = 'Building G'
                elif first_char == 'J':
                    building_display = 'Building J'
                else:
                    building_display = str(row.get('building_name') or row.get('Building', ''))
            else:
                building_display = str(row.get('building_name') or row.get('Building', ''))
            
            # Inspection Date - FIRST COLUMN
            insp_date = row.get('inspection_date') or row.get('InspectionDate')
            if pd.notna(insp_date):
                if isinstance(insp_date, str):
                    ws.write(row_idx, 0, insp_date, base_fmt)
                else:
                    ws.write(row_idx, 0, insp_date, date_row_fmt)
            else:
                ws.write(row_idx, 0, '', base_fmt)
            
            # Write data (NO Issue Description column!)
            ws.write(row_idx, 1, building_display, base_fmt)  # Building
            ws.write(row_idx, 2, unit, base_fmt)  # Unit
            ws.write(row_idx, 3, str(row.get('Room', '')), base_fmt)
            ws.write(row_idx, 4, str(row.get('Component', '')), base_fmt)
            ws.write(row_idx, 5, str(row.get('Trade', '')), base_fmt)
            ws.write(row_idx, 6, str(row.get('Urgency', '')), base_fmt)
            ws.write(row_idx, 7, str(row.get('StatusClass', '')), base_fmt)
            ws.write(row_idx, 8, str(row.get('InspectorNotes', '')), notes_fmt)
            
            # Photo column (J) - Leave blank, images will be embedded in Pass 2
            ws.write(row_idx, 9, '', base_fmt)
        
        return sheet_name
    
    def _create_settlement_sheet(self, workbook, metrics, formats):
        """Create settlement readiness sheet"""
        ws = workbook.add_worksheet("🏠 Settlement Readiness")
        ws.set_column('A:A', 40)
        ws.set_column('B:B', 15)
        ws.set_column('C:C', 15)
        ws.set_column('D:D', 20)
        
        # Headers
        headers = ['Category', 'Units', 'Percentage', 'Criteria']
        for col_idx, header in enumerate(headers):
            ws.write(0, col_idx, header, formats['header'])
        
        # Data
        settlement_data = [
            ('✅ Minor Work Required', metrics['ready_units'], f"{metrics['ready_pct']:.1f}%", '0-2 defects', formats['ready']),
            ('⚠️ Intermediate Remediation', metrics['minor_work_units'], f"{metrics['minor_pct']:.1f}%", '3-7 defects', formats['minor']),
            ('🔧 Major Work Required', metrics['major_work_units'], f"{metrics['major_pct']:.1f}%", '8-15 defects', formats['major']),
            ('🚧 Extensive Work Required', metrics['extensive_work_units'], f"{metrics['extensive_pct']:.1f}%", '15+ defects', formats['extensive'])
        ]
        
        for row_idx, (category, units, percentage, criteria, fmt) in enumerate(settlement_data, 1):
            ws.write(row_idx, 0, category, fmt)
            ws.write(row_idx, 1, units, fmt)
            ws.write(row_idx, 2, percentage, fmt)
            ws.write(row_idx, 3, criteria, fmt)
    
    def _create_summary_sheet(self, workbook, summary_df, sheet_name, formats):
        """Create summary sheet (Trade/Room/Component/Unit)"""
        ws = workbook.add_worksheet(sheet_name)
        ws.set_column('A:A', 40)
        ws.set_column('B:B', 15)
        
        # Headers
        headers = list(summary_df.columns)
        for col_idx, header in enumerate(headers):
            ws.write(0, col_idx, header, formats['header'])
        
        # Data
        for row_idx, (_, row) in enumerate(summary_df.iterrows(), start=1):
            is_alt = (row_idx % 2 == 0)
            fmt = formats['cell_alt'] if is_alt else formats['cell']
            
            for col_idx, value in enumerate(row):
                ws.write(row_idx, col_idx, value, fmt)
    
    def _create_all_inspections_sheet(self, workbook, data_df, metrics, formats):
        """
        Create All Inspections sheet - shows defect items without photos
        Format: Inspection Date | Building | Unit | Room | Component | Status | Trade | Priority | Inspector Notes
        """
        ws = workbook.add_worksheet("📝 All Inspections")
        
        # Column widths
        ws.set_column('A:A', 15)  # Inspection Date
        ws.set_column('B:B', 15)  # Building
        ws.set_column('C:C', 12)  # Unit
        ws.set_column('D:D', 20)  # Room
        ws.set_column('E:E', 25)  # Component
        ws.set_column('F:F', 12)  # Status
        ws.set_column('G:G', 15)  # Trade
        ws.set_column('H:H', 12)  # Priority
        ws.set_column('I:I', 40)  # Inspector Notes
        
        # Headers - Inspection Date FIRST, no Description
        headers = ['Inspection Date', 'Building', 'Unit', 'Room', 'Component', 'Status', 'Trade', 'Priority', 'Inspector Notes']
        for col_idx, header in enumerate(headers):
            ws.write(0, col_idx, header, formats['header'])
        
        # Date format
        date_fmt = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'num_format': 'dd/mm/yyyy'
        })
        date_fmt_alt = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'num_format': 'dd/mm/yyyy',
            'bg_color': '#F7F9FC'
        })
        
        # Data rows
        for row_idx, (_, row) in enumerate(data_df.iterrows(), start=1):
            is_alt = (row_idx % 2 == 0)
            base_fmt = formats['cell_alt'] if is_alt else formats['cell']
            notes_fmt = formats['notes_alt'] if is_alt else formats['notes']
            date_row_fmt = date_fmt_alt if is_alt else date_fmt
            
            # Get unit for building determination
            unit = str(row.get('unit') or row.get('Unit', ''))
            
            # Building - Simplified based on unit prefix
            if unit and len(unit) > 0:
                first_char = unit[0].upper()
                if first_char == 'G':
                    building_display = 'Building G'
                elif first_char == 'J':
                    building_display = 'Building J'
                else:
                    building_display = str(row.get('building_name') or row.get('Building', ''))
            else:
                building_display = str(row.get('building_name') or row.get('Building', ''))
            
            # Inspection Date - FIRST COLUMN
            insp_date = row.get('inspection_date') or row.get('InspectionDate')
            if pd.notna(insp_date):
                if isinstance(insp_date, str):
                    ws.write(row_idx, 0, insp_date, base_fmt)
                else:
                    ws.write(row_idx, 0, insp_date, date_row_fmt)
            else:
                ws.write(row_idx, 0, '', base_fmt)
            
            # Write remaining data (no Description column!)
            ws.write(row_idx, 1, building_display, base_fmt)
            ws.write(row_idx, 2, unit, base_fmt)
            ws.write(row_idx, 3, str(row.get('Room', '')), base_fmt)
            ws.write(row_idx, 4, str(row.get('Component', '')), base_fmt)
            ws.write(row_idx, 5, str(row.get('StatusClass', '')), base_fmt)
            ws.write(row_idx, 6, str(row.get('Trade', '')), base_fmt)
            ws.write(row_idx, 7, str(row.get('Urgency', '')), base_fmt)
            ws.write(row_idx, 8, str(row.get('InspectorNotes', '')), notes_fmt)
    
    def _create_component_details_sheet(self, workbook, data_df, formats):
        """
        Create Component Details sheet - shows Trade, Room, Component, and Units with Defects
        Format matches template: Trade | Room | Component | Units with Defects
        """
        ws = workbook.add_worksheet("🔍 Component Details")
        
        # Column widths
        ws.set_column('A:A', 25)  # Trade
        ws.set_column('B:B', 30)  # Room
        ws.set_column('C:C', 30)  # Component
        ws.set_column('D:D', 100)  # Units with Defects (wide for unit list)
        
        # Headers
        headers = ['Trade', 'Room', 'Component', 'Units with Defects']
        for col_idx, header in enumerate(headers):
            ws.write(0, col_idx, header, formats['header'])
        
        # Calculate component stats with units
        try:
            if 'Component' in data_df.columns and 'Trade' in data_df.columns and len(data_df) > 0:
                component_stats = []
                
                # Group by Trade, Room, Component
                for trade in data_df['Trade'].dropna().unique():
                    trade_items = data_df[data_df['Trade'] == trade]
                    
                    for room in trade_items['Room'].dropna().unique():
                        room_items = trade_items[trade_items['Room'] == room]
                        
                        for component in room_items['Component'].dropna().unique():
                            if not component:  # Skip empty components
                                continue
                            
                            component_items = room_items[room_items['Component'] == component]
                            
                            # Get list of units with this defect
                            units_with_defects = []
                            for _, item in component_items.iterrows():
                                unit = str(item.get('unit') or item.get('Unit', '')).strip()
                                if unit and unit not in units_with_defects:
                                    units_with_defects.append(unit)
                            
                            # Sort units
                            units_with_defects.sort()
                            
                            component_stats.append({
                                'Trade': str(trade),
                                'Room': str(room),
                                'Component': str(component),
                                'Units': ', '.join(units_with_defects),
                                'DefectCount': len(component_items)
                            })
                
                # Sort by trade, then by defect count descending
                component_stats.sort(key=lambda x: (x['Trade'], -x['DefectCount']))
                
                # Write data
                for row_idx, stat in enumerate(component_stats, start=1):
                    is_alt = (row_idx % 2 == 0)
                    base_fmt = formats['cell_alt'] if is_alt else formats['cell']
                    
                    ws.write(row_idx, 0, stat['Trade'], base_fmt)
                    ws.write(row_idx, 1, stat['Room'], base_fmt)
                    ws.write(row_idx, 2, stat['Component'], base_fmt)
                    ws.write(row_idx, 3, stat['Units'], base_fmt)
            else:
                # No data
                ws.write(1, 0, 'No component data available', formats['cell'])
        except Exception as e:
            logger.error(f"Error creating component details: {str(e)}")
            ws.write(1, 0, f'Error generating component details: {str(e)}', formats['cell'])
    
    def _create_metadata_sheet(self, workbook, metrics, formats):
        """Create metadata sheet"""
        ws = workbook.add_worksheet("📄 Report Metadata")
        ws.set_column('A:A', 30)
        ws.set_column('B:B', 40)
        
        quality_score = max(0, 100 - metrics.get('defect_rate', 0))
        
        melbourne_tz = pytz.timezone('Australia/Melbourne')
        melbourne_time = datetime.now(melbourne_tz)
        
        metadata = [
            ('Report Generated', melbourne_time.strftime('%Y-%m-%d %H:%M:%S AEDT')),
            ('Report Version', '3.0 Professional with Photos'),
            ('Building Name', metrics['building_name']),
            ('Total Units', str(metrics['total_units'])),
            ('Total Defects', str(metrics['total_defects'])),
            ('Quality Score', f"{quality_score:.1f}/100"),
            ('Photos Included', str(metrics.get('photo_count', 0))),
            ('Inspector Notes', str(metrics.get('notes_count', 0))),
            ('Data Source', 'SafetyCulture API via PostgreSQL'),
            ('Processing Engine', 'Professional Report Generator with Photo Support')
        ]
        
        # Headers
        ws.write(0, 0, 'Property', formats['header'])
        ws.write(0, 1, 'Value', formats['header'])
        
        # Data
        for row_idx, (prop, value) in enumerate(metadata, 1):
            ws.write(row_idx, 0, prop, formats['label'])
            ws.write(row_idx, 1, value, formats['cell'])
    
    def _create_workflow_tracker_sheet(self, workbook, data_df, metrics, formats):
        """
        Create Workflow Tracker sheet with status dates and color coding
        This is the enhanced format requested by the user!
        """
        ws = workbook.add_worksheet("🔄 Workflow Tracker")
        
        # Column widths
        ws.set_column('A:A', 12)  # Date
        ws.set_column('B:B', 15)  # Inspection ID
        ws.set_column('C:C', 12)  # Apartment #
        ws.set_column('D:D', 18)  # Building
        ws.set_column('E:E', 20)  # Room
        ws.set_column('F:F', 35)  # Defect Description
        ws.set_column('G:G', 18)  # Trade
        ws.set_column('H:H', 15)  # Photo Link
        ws.set_column('I:I', 15)  # Sent to Builder
        ws.set_column('J:J', 15)  # Builder Done
        ws.set_column('K:K', 15)  # Owner Confirmed
        
        # Headers
        headers = [
            'Date', 'Inspection', 'Apt #', 'Building', 'Room', 
            'Defect', 'Trade', 'Photo', '📤 Sent to Builder', 
            '✅ Builder Done', '👍 Owner Confirmed'
        ]
        for col_idx, header in enumerate(headers):
            ws.write(0, col_idx, header, formats['header'])
        
        # Add color formats for status
        yellow_fmt = workbook.add_format({
            'bg_color': '#FFF3C4',
            'border': 1,
            'align': 'left',
            'valign': 'vcenter'
        })
        blue_fmt = workbook.add_format({
            'bg_color': '#DAEEF3',
            'border': 1,
            'align': 'left',
            'valign': 'vcenter'
        })
        green_fmt = workbook.add_format({
            'bg_color': '#C8E6C9',
            'border': 1,
            'align': 'left',
            'valign': 'vcenter'
        })
        date_fmt = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'num_format': 'dd/mm/yyyy'
        })
        
        # Data rows
        for row_idx, (_, row) in enumerate(data_df.iterrows(), start=1):
            # Determine row color based on workflow status
            # Only check for owner confirmation - that's the only real completion marker
            has_owner = pd.notna(row.get('owner_signoff_timestamp'))
            
            # Choose background color
            if has_owner:
                row_fmt = green_fmt  # 🟢 Complete (owner confirmed)
            else:
                row_fmt = yellow_fmt # 🟡 Pending (everything else - sent to builder but not done)
            
            # Write data with color coding
            # Date
            insp_date = row.get('inspection_date') or row.get('InspectionDate')
            if pd.notna(insp_date):
                if isinstance(insp_date, str):
                    ws.write(row_idx, 0, insp_date, row_fmt)
                else:
                    ws.write(row_idx, 0, insp_date, date_fmt)
            else:
                ws.write(row_idx, 0, '', row_fmt)
            
            # Inspection ID (first 8 chars)
            inspection_id = str(row.get('inspection_id', ''))[:8]
            ws.write(row_idx, 1, inspection_id, row_fmt)
            
            # Apartment # (Unit)
            unit = str(row.get('unit') or row.get('Unit', ''))
            ws.write(row_idx, 2, unit, row_fmt)
            
            # Building - Simplified based on unit prefix
            # If unit starts with G → "Building G", if starts with J → "Building J"
            if unit and len(unit) > 0:
                first_char = unit[0].upper()
                if first_char == 'G':
                    building_display = 'Building G'
                elif first_char == 'J':
                    building_display = 'Building J'
                else:
                    # Fallback to full building name if not G or J
                    building_display = str(row.get('building_name') or row.get('Building', ''))
            else:
                building_display = str(row.get('building_name') or row.get('Building', ''))
            
            ws.write(row_idx, 3, building_display, row_fmt)
            
            # Room
            ws.write(row_idx, 4, str(row.get('Room') or row.get('room', '')), row_fmt)
            
            # Defect Description
            defect_desc = row.get('IssueDescription') or row.get('description') or row.get('Component', '')
            ws.write(row_idx, 5, str(defect_desc), row_fmt)
            
            # Trade
            ws.write(row_idx, 6, str(row.get('Trade') or row.get('trade', '')), row_fmt)
            
            # Photo Link - Don't use API URL (requires authentication)
            # Instead, refer to Defects Only sheet where photos are embedded
            photo_url = row.get('photo_url')
            if pd.notna(photo_url) and photo_url:
                # Don't write URL - it won't work without API key
                # Just indicate photo is available in other sheet
                ws.write(row_idx, 7, 'See Defects Only', row_fmt)
            else:
                ws.write(row_idx, 7, 'No photo', row_fmt)
            
            # Sent to Builder date (when defect was created)
            sent_date = row.get('created_at')
            if pd.notna(sent_date):
                if isinstance(sent_date, str):
                    ws.write(row_idx, 8, sent_date, row_fmt)
                else:
                    ws.write(row_idx, 8, sent_date, date_fmt)
            else:
                ws.write(row_idx, 8, '', row_fmt)
            
            # Builder Done date - BLANK (to be filled manually when work complete)
            # Note: planned_completion is just a plan, not actual completion!
            ws.write(row_idx, 9, '', row_fmt)
            
            # Owner Confirmed date - Use owner_signoff_timestamp if exists
            # This is the only one that's actually populated when confirmed
            owner_date = row.get('owner_signoff_timestamp')
            if pd.notna(owner_date):
                if isinstance(owner_date, str):
                    ws.write(row_idx, 10, owner_date, row_fmt)
                else:
                    ws.write(row_idx, 10, owner_date, date_fmt)
            else:
                ws.write(row_idx, 10, '', row_fmt)
        
        # Add legend at the bottom
        legend_row = len(data_df) + 3
        ws.write(legend_row, 0, 'Legend:', formats['label'])
        ws.write(legend_row + 1, 0, '🟡 Yellow = Sent to Builder (Pending - default)', yellow_fmt)
        ws.write(legend_row + 2, 0, '🔵 Blue = Builder Complete (Fill date when work done)', blue_fmt)
        ws.write(legend_row + 3, 0, '🟢 Green = Owner Confirmed (Automatically from signoff)', green_fmt)
        
        # Add note about manual updates
        ws.write(legend_row + 5, 0, 'NOTE: Fill "Builder Done" date manually when builder completes work', formats['label'])
        ws.write(legend_row + 6, 0, '      Row will turn blue automatically when date is entered!', formats['label'])
        
        # Add conditional formatting: Turn rows blue when Builder Done (column J) has value
        # This makes the row turn blue automatically when someone fills in the date
        if len(data_df) > 0:
            # Apply to all data rows (row 2 onwards, since row 1 is header)
            data_rows = f'A2:K{len(data_df) + 1}'
            
            # Rule 1: Green if Owner Confirmed has value (column K)
            ws.conditional_format(data_rows, {
                'type': 'formula',
                'criteria': '=$K2<>""',
                'format': green_fmt
            })
            
            # Rule 2: Blue if Builder Done has value but Owner Confirmed doesn't (column J has value, K blank)
            ws.conditional_format(data_rows, {
                'type': 'formula', 
                'criteria': '=AND($J2<>"",$K2="")',
                'format': blue_fmt
            })
            
            # Yellow is the default background color already applied
    
    def _add_photos_with_openpyxl(self, excel_path: str, data_df: pd.DataFrame):
        """
        Pass 2: Add photos to Excel using openpyxl
        
        Args:
            excel_path: Path to Excel file created by xlsxwriter
            data_df: DataFrame with photo URLs
        """
        try:
            import openpyxl
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.utils import get_column_letter
            
            logger.info(f"Opening Excel file with openpyxl: {excel_path}")
            
            # Load workbook
            wb = openpyxl.load_workbook(excel_path)
            
            # Find the "Defects Only" sheet (renamed from All Defects)
            sheet_name = "⚠️ Defects Only"
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found in workbook")
                wb.save(excel_path)
                return
            
            ws = wb[sheet_name]
            logger.info(f"Found sheet: {sheet_name}")
            
            # Column J is for photos (column 10, 1-indexed) - NOW THAT WE ADDED BUILDING & UNIT
            photo_col = 10
            photo_col_letter = get_column_letter(photo_col)
            
            # Set column width for photos
            ws.column_dimensions[photo_col_letter].width = 20
            
            # Iterate through data rows (skip header row 0)
            photos_added = 0
            photos_failed = 0
            
            for row_idx, (_, row) in enumerate(data_df.iterrows(), start=2):  # Excel rows start at 1, +1 for header
                try:
                    photo_url = row.get('photo_url')
                    
                    if pd.notna(photo_url) and photo_url:
                        # Download photo
                        img = self.download_photo(photo_url)
                        
                        if img:
                            # Resize to thumbnail
                            img_bytes = self.resize_to_thumbnail(img, size=(150, 150))
                            
                            # Create openpyxl image object
                            xl_img = XLImage(img_bytes)
                            
                            # Set image size (in pixels)
                            xl_img.width = 150
                            xl_img.height = 150
                            
                            # Position image in cell (column H, current row)
                            cell_ref = f'{photo_col_letter}{row_idx}'
                            xl_img.anchor = cell_ref
                            
                            # Add image to worksheet
                            ws.add_image(xl_img)
                            
                            # Set row height (in points, 120 points ≈ 160 pixels)
                            ws.row_dimensions[row_idx].height = 120
                            
                            photos_added += 1
                            
                            if photos_added % 10 == 0:
                                logger.info(f"Added {photos_added} photos...")
                        else:
                            photos_failed += 1
                            # Write fallback text
                            ws.cell(row=row_idx, column=photo_col, value="Photo unavailable")
                    else:
                        # No photo for this row
                        ws.row_dimensions[row_idx].height = 30
                        
                except Exception as e:
                    logger.error(f"Error adding photo for row {row_idx}: {str(e)}")
                    photos_failed += 1
                    continue
            
            # Save workbook
            wb.save(excel_path)
            logger.info(f"Photos added: {photos_added}, Failed: {photos_failed}")
            logger.info(f"Excel file with photos saved: {excel_path}")
            
        except Exception as e:
            logger.error(f"Error in openpyxl photo pass: {str(e)}")
            import traceback
            traceback.print_exc()
            # Don't raise - file should still have all the data without photos


def create_professional_excel_from_database(
    inspection_ids: List[int],
    db_connection,
    api_key: str,
    output_path: str,
    report_type: str = "single"
) -> bool:
    """
    Main function to create professional Excel report from database
    
    Args:
        inspection_ids: List of inspection IDs
        db_connection: Database connection
        api_key: SafetyCulture API key
        output_path: Output file path
        report_type: "single" or "multi"
        
    Returns:
        True if successful
    """
    try:
        generator = ProfessionalExcelGeneratorAPI(api_key)
        
        if report_type == "single" and len(inspection_ids) == 1:
            # Query single inspection
            inspection_data, defects = _query_inspection_data(db_connection, inspection_ids[0])
            return generator.generate_professional_report(inspection_data, defects, output_path)
        
        elif report_type == "multi":
            # Multi-inspection support (e.g., building report with multiple units)
            logger.info(f"Creating multi-inspection report for {len(inspection_ids)} inspections")
            
            # Query all inspections and combine defects
            all_defects = []
            building_name = None
            address = None
            inspection_dates = []
            
            for inspection_id in inspection_ids:
                try:
                    inspection_data, defects = _query_inspection_data(db_connection, inspection_id)
                    
                    # Collect building info from first inspection
                    if building_name is None:
                        building_name = inspection_data.get('building_name', 'Building')
                        address = inspection_data.get('address', 'Address')
                    
                    # Collect inspection dates
                    if inspection_data.get('inspection_date'):
                        inspection_dates.append(inspection_data['inspection_date'])
                    
                    # Add all defects
                    all_defects.extend(defects)
                    
                except Exception as e:
                    logger.error(f"Error querying inspection {inspection_id}: {str(e)}")
                    continue
            
            if len(all_defects) == 0:
                logger.warning("No defects found across all inspections")
                return False
            
            # Create combined inspection data
            combined_inspection_data = {
                'id': 'multi',
                'inspection_date': max(inspection_dates) if inspection_dates else 'N/A',
                'inspector_name': 'Multiple Inspectors',
                'total_defects': len(all_defects),
                'building_name': building_name,
                'address': address,
                'unit': 'Multiple Units',
                'unit_type': 'Mixed',
                'total_items': sum(1 for _ in all_defects)  # Approximate
            }
            
            logger.info(f"Combined {len(all_defects)} defects from {len(inspection_ids)} inspections")
            
            # Generate report with combined data
            return generator.generate_professional_report(combined_inspection_data, all_defects, output_path)
        
        else:
            logger.error(f"Invalid report type: {report_type}")
            return False
            
    except Exception as e:
        logger.error(f"Error creating professional Excel report: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def _query_inspection_data(db_connection, inspection_id: int) -> tuple:
    """
    Query inspection data from database
    
    Args:
        db_connection: Database connection
        inspection_id: Inspection ID
        
    Returns:
        tuple: (inspection_data dict, defects list)
    """
    logger.info(f"Querying inspection data for ID: {inspection_id} (type: {type(inspection_id)})")
    
    cursor = db_connection.cursor()
    
    try:
        # Query inspection metadata (no unit columns in inspections table)
        cursor.execute("""
            SELECT i.id, i.inspection_date, i.inspector_name, i.total_defects,
                   b.name as building_name, b.address
            FROM inspector_inspections i
            JOIN inspector_buildings b ON i.building_id = b.id
            WHERE i.id = %s
        """, (inspection_id,))
        
        row = cursor.fetchone()
        if not row:
            logger.error(f"Inspection {inspection_id} not found in database")
            raise ValueError(f"Inspection {inspection_id} not found")
        
        inspection_data = {
            'id': row[0],
            'inspection_date': row[1].strftime('%Y-%m-%d') if row[1] else 'N/A',
            'inspector_name': row[2] or 'N/A',
            'total_defects': row[3] or 0,
            'building_name': row[4] or 'Building',
            'address': row[5] or 'Address',
            'unit': 'Multiple Units',
            'unit_type': 'Mixed'
        }
        
        # Get unit info from first inspection item
        cursor.execute("""
            SELECT unit, unit_type
            FROM inspector_inspection_items
            WHERE inspection_id = %s
            LIMIT 1
        """, (inspection_id,))
        
        unit_row = cursor.fetchone()
        if unit_row and len(unit_row) >= 2:
            inspection_data['unit'] = unit_row[0] or 'Unit'
            inspection_data['unit_type'] = unit_row[1] or 'Apartment'
        else:
            logger.warning(f"No inspection items found for inspection {inspection_id}")
    
    except Exception as e:
        logger.error(f"Error querying inspection metadata: {e}")
        cursor.close()
        raise
    
    # Query defects with photos, dates, and building info (case-insensitive status filter)
    cursor.execute("""
        SELECT ii.room, ii.component, ii.notes, ii.trade, ii.urgency, ii.status_class,
            ii.photo_url, ii.photo_media_id, ii.inspector_notes,
            ii.inspection_date, ii.created_at, ii.planned_completion, ii.owner_signoff_timestamp,
            ii.unit, b.name as building_name
        FROM inspector_inspection_items ii
        JOIN inspector_inspections i ON ii.inspection_id = i.id
        JOIN inspector_buildings b ON i.building_id = b.id
        WHERE ii.inspection_id = %s
        AND LOWER(ii.status_class) = 'not ok'
        ORDER BY ii.room, ii.component
    """, (inspection_id,))
    
    defects = []
    for row in cursor.fetchall():
        defects.append({
            'room': row[0],
            'component': row[1],
            'description': row[2],
            'trade': row[3],
            'priority': row[4],
            'status': row[5],
            'photo_url': row[6],
            'photo_media_id': row[7],
            'inspector_notes': row[8],
            'inspection_date': row[9],
            'created_at': row[10],
            'planned_completion': row[11],
            'owner_signoff_timestamp': row[12],
            'unit': row[13],
            'building_name': row[14]
        })
    
    logger.info(f"Found {len(defects)} defects for inspection {inspection_id}")
    
    if len(defects) == 0:
        logger.warning(f"No defects found for inspection {inspection_id} - check status_class values")
    
    cursor.close()
    return inspection_data, defects


if __name__ == "__main__":
    print("Professional Excel Generator API - Ready!")
    print("Combines template dashboard with API photo support")