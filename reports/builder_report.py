"""
Builder Defect Management Report Generator - COMPLETE VERSION
=============================================================
"""

import pandas as pd
from datetime import datetime
from io import BytesIO
import streamlit as st
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import logging
import sqlite3
import os
import tempfile

from reports.report_utils import (
    ReportStyler, 
    ReportDataProcessor, 
    ReportMetadata,
    create_summary_dataframe
)

logger = logging.getLogger(__name__)


class BuilderReportGenerator:
    """Generate comprehensive defect management reports for builders"""
    
    def __init__(self, db_manager):
        self.db_manager = db_manager
        self.db_path = db_manager.db_path if hasattr(db_manager, 'db_path') else "building_inspection.db"
        self.styler = ReportStyler()
        self.processor = ReportDataProcessor()
    
    def _get_fresh_connection(self):
        """Always get a fresh database connection"""
        return sqlite3.connect(self.db_path, check_same_thread=False, detect_types=0)
    
    def generate_excel_report(self, builder_name=None, inspection_id=None, include_photos=False, status_filter=None, include_files_sheet=False):
        """Generate Excel report with multiple sheets
        
        Args:
            include_files_sheet: If True, include the Files sheet (for ZIP packages).
                                 If False, skip Files sheet (for standalone Excel downloads).
        """
        conn = None
        temp_files = []  # Track temp files to clean up
        try:
            # Fetch data with fresh connections
            work_orders_df = self._get_work_orders(builder_name, inspection_id, status_filter)
            defects_df = self._get_defects(inspection_id)
            
            if work_orders_df.empty:
                st.warning("⚠️ No work orders found for the selected criteria")
                return None
            
            logger.info(f"Generating Excel report: {len(work_orders_df)} work orders, {len(defects_df)} defects")
            
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)
            
            # Add metadata sheet
            metadata_info = {
                'title': 'Builder Defect Management Report',
                'generated_by': builder_name or 'All Builders',
                'report_type': 'Defect Analysis',
                'period': datetime.now().strftime('%Y-%m-%d'),
                'total_records': len(work_orders_df),
                'filters': f"Builder: {builder_name or 'All'}, Inspection: {inspection_id or 'All'}"
            }
            ReportMetadata.create_metadata_sheet(wb, metadata_info)
            
            # Create sheets
            self._create_summary_sheet(wb, work_orders_df, defects_df)
            self._create_work_orders_sheet(wb, work_orders_df)
            self._create_by_trade_sheet(wb, defects_df)
            self._create_by_unit_sheet(wb, defects_df)
            self._create_progress_sheet(wb, work_orders_df)
            
            if include_photos:
                temp_files = self._create_photos_sheet(wb, defects_df)
                # Only include Files sheet if explicitly requested (for ZIP packages)
                if include_files_sheet:
                    self._create_files_sheet(wb, defects_df)
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # ✅ NOW clean up temp files AFTER workbook is saved
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                        logger.info(f"🧹 Cleaned up temp file: {temp_file}")
                except Exception as e:
                    logger.warning(f"Could not delete temp file {temp_file}: {e}")
            
            logger.info("✅ Excel report generated successfully")
            return output
            
        except Exception as e:
            logger.error(f"❌ Error generating Excel report: {e}")
            st.error(f"Error generating report: {e}")
            import traceback
            logger.error(traceback.format_exc())
            
            # Clean up temp files even on error
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                except:
                    pass
            
            return None
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass
    
    def _get_work_orders(self, builder_name=None, inspection_id=None, status_filter=None):
        """Fetch work orders - uses fresh connection with CORRECT column names"""
        conn = None
        try:
            conn = self._get_fresh_connection()
            
            # Use ACTUAL column names from inspector_work_orders table
            query = """
                SELECT 
                    wo.id,
                    wo.id as work_order_number,
                    wo.unit,
                    wo.trade,
                    wo.component,
                    wo.room,
                    wo.urgency as priority,
                    wo.status,
                    wo.assigned_to,
                    wo.notes as description,
                    wo.created_at,
                    wo.updated_at,
                    wo.planned_date as due_date,
                    wo.started_date,
                    wo.completed_date as completed_at,
                    wo.builder_notes,
                    wo.notes,
                    i.inspection_date,
                    b.name as building_name
                FROM inspector_work_orders wo
                LEFT JOIN inspector_inspections i ON wo.inspection_id = i.id
                LEFT JOIN inspector_buildings b ON i.building_id = b.id
                WHERE 1=1
            """
            
            params = []
            
            if inspection_id:
                query += " AND wo.inspection_id = ?"
                params.append(inspection_id)
            
            # Add status filter if provided
            if status_filter:
                query += f" AND {status_filter}"
                logger.info(f"Applying status filter: {status_filter}")
            
            # Order by status priority, then by date
            query += """
                ORDER BY 
                    CASE wo.status 
                        WHEN 'pending' THEN 1 
                        WHEN 'in_progress' THEN 2 
                        WHEN 'waiting_approval' THEN 3 
                        WHEN 'approved' THEN 4 
                        ELSE 5 
                    END,
                    wo.updated_at DESC
            """
            
            logger.info(f"Executing query with params: {params}")
            df = pd.read_sql_query(query, conn, params=params if params else None)
            logger.info(f"Query returned {len(df)} rows")
            
            # Create a better title from unit + room + component
            if 'unit' in df.columns and 'room' in df.columns and 'component' in df.columns:
                df['title'] = df.apply(
                    lambda row: f"Unit {row['unit']} - {row['room']} - {row['component']}", 
                    axis=1
                )
            
            return df
            
        except Exception as e:
            logger.error(f"Error fetching work orders: {e}")
            return pd.DataFrame()
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass
    
    def _get_defects(self, inspection_id=None):
        """Fetch defects - uses fresh connection"""
        conn = None
        try:
            conn = self._get_fresh_connection()
            
            query = """
                SELECT 
                    id,
                    inspection_id,
                    unit,
                    unit_type,
                    room,
                    component,
                    trade,
                    status_class as status,
                    urgency,
                    planned_completion,
                    created_at
                FROM inspector_inspection_items
                WHERE status_class = 'Not OK'
            """
            
            params = []
            if inspection_id:
                query += " AND inspection_id = ?"
                params.append(inspection_id)
            
            df = pd.read_sql_query(query, conn, params=params if params else None)
            return df
            
        except Exception as e:
            logger.error(f"Error fetching defects: {e}")
            return pd.DataFrame()
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass
    
    def _create_summary_sheet(self, wb, work_orders_df, defects_df):
        """Create executive summary sheet - FIXED STATUS VALUES"""
        ws = wb.create_sheet("📊 Executive Summary")
        
        total_wo = len(work_orders_df)
        
        # ✅ FIXED: Use correct status values
        pending = len(work_orders_df[work_orders_df['status'] == 'pending'])
        in_progress = len(work_orders_df[work_orders_df['status'] == 'in_progress'])
        waiting = len(work_orders_df[work_orders_df['status'] == 'waiting_approval'])
        approved = len(work_orders_df[work_orders_df['status'] == 'approved'])
        
        # Calculate completion rate (approved / total)
        completion_rate = self.processor.calculate_completion_rate(approved, total_wo)
        
        total_defects = len(defects_df)
        urgent = len(defects_df[defects_df['urgency'] == 'Urgent']) if len(defects_df) > 0 else 0
        high_priority = len(defects_df[defects_df['urgency'] == 'High Priority']) if len(defects_df) > 0 else 0
        
        summary_data = {
            'Total Work Orders': total_wo,
            'Pending Work Orders': pending,
            'In Progress': in_progress,
            'Awaiting Approval': waiting,
            'Approved Work Orders': approved,
            'Completion Rate %': completion_rate,
            '': '',
            'Total Defects': total_defects,
            'Urgent Defects': urgent,
            'High Priority Defects': high_priority,
            'Normal Priority': total_defects - urgent - high_priority,
        }
        
        if len(defects_df) > 0:
            summary_data[''] = ''
            most_common_trade = defects_df['trade'].mode()
            summary_data['Most Common Trade'] = most_common_trade[0] if len(most_common_trade) > 0 else 'N/A'
            summary_data['Unit with Most Defects'] = defects_df.groupby('unit').size().idxmax()
            summary_data['Avg Defects per Unit'] = f"{len(defects_df) / defects_df['unit'].nunique():.1f}"
        
        summary_df = create_summary_dataframe(summary_data)
        
        for idx, row in enumerate(summary_df.itertuples(index=False), start=1):
            ws.cell(row=idx, column=1, value=row.Metric).font = Font(bold=True, size=11)
            ws.cell(row=idx, column=2, value=row.Value).font = Font(size=11)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        
        ws.insert_rows(1)
        ws['A1'] = 'BUILDER DEFECT MANAGEMENT - EXECUTIVE SUMMARY'
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.styler.HEADER_COLOR, end_color=self.styler.HEADER_COLOR, fill_type="solid")
        ws.merge_cells('A1:B1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    def _create_work_orders_sheet(self, wb, work_orders_df):
        """Create work orders details sheet with full notes"""
        ws = wb.create_sheet("📋 Work Orders")
        
        # DEBUG: Log what columns we actually have
        logger.info(f"Available columns: {work_orders_df.columns.tolist()}")
        if len(work_orders_df) > 0:
            sample_row = work_orders_df.iloc[0]
            logger.info(f"Sample row timestamps - created_at: {sample_row.get('created_at')}, updated_at: {sample_row.get('updated_at')}, started_date: {sample_row.get('started_date')}, completed_at: {sample_row.get('completed_at')}")
            logger.info(f"Sample row status: {sample_row.get('status')}")
        
        # Create a copy to avoid modifying the original
        df_copy = work_orders_df.copy()
        
        # ✅ IMPROVED: Better last activity logic with proper datetime handling
        def get_last_activity(row):
            status = row.get('status', '')
            
            # Define priority order based on status
            if status in ['waiting_approval', 'approved']:
                # Completed items: prioritize completed_at
                fields_to_check = ['completed_at', 'updated_at', 'started_date', 'created_at']
            elif status == 'in_progress':
                # In progress: prioritize started_date, then updated_at
                fields_to_check = ['started_date', 'updated_at', 'created_at']
            else:
                # Pending or other: prioritize updated_at, then created_at
                fields_to_check = ['updated_at', 'created_at']
            
            # Check each field in priority order
            for field in fields_to_check:
                value = row.get(field)
                if pd.notna(value) and value not in ['', None, 'None', 'NaT']:
                    # Try to parse and return as string
                    try:
                        # If it's already a datetime
                        if isinstance(value, pd.Timestamp):
                            return value
                        # Try parsing string
                        parsed = pd.to_datetime(value, errors='coerce')
                        if pd.notna(parsed):
                            return parsed
                    except:
                        pass
            
            # If nothing found, return None
            return None
        
        df_copy['last_activity'] = df_copy.apply(get_last_activity, axis=1)
        
        # DEBUG: Check what we got
        logger.info(f"Last activity sample: {df_copy['last_activity'].head(3).tolist()}")
        logger.info(f"Last activity types: {df_copy['last_activity'].apply(type).unique()}")
        
        # Use only columns that exist
        available_columns = []
        column_mapping = {}
        
        if 'work_order_number' in df_copy.columns:
            available_columns.append('work_order_number')
            column_mapping['work_order_number'] = 'WO Number'
        elif 'id' in df_copy.columns:
            available_columns.append('id')
            column_mapping['id'] = 'WO ID'
        
        # Build columns list
        potential_columns = [
            ('title', 'Title'),
            ('unit', 'Unit'),
            ('room', 'Room'),
            ('component', 'Component'),
            ('trade', 'Trade'),
            ('priority', 'Priority'),
            ('status', 'Status'),
            ('assigned_to', 'Assigned To'),
            ('building_name', 'Building'),
            ('created_at', 'Created'),
            ('last_activity', 'Last Activity'),
            ('planned_date', 'Target Date'),
            ('builder_notes', 'Work History & Notes')
        ]
        
        for col, display_name in potential_columns:
            if col in df_copy.columns:
                available_columns.append(col)
                column_mapping[col] = display_name
        
        if not available_columns:
            ws['A1'] = 'No data available'
            return
        
        display_df = df_copy[available_columns].copy()
        display_df.columns = [column_mapping[col] for col in available_columns]
        
        # ✅ IMPROVED: Better datetime formatting with proper error handling
        datetime_cols = ['Created', 'Last Activity']
        for col in datetime_cols:
            if col in display_df.columns:
                def format_datetime(x):
                    if pd.isna(x) or x in ['', None, 'None', 'NaT']:
                        return ''
                    try:
                        if isinstance(x, pd.Timestamp):
                            return x.strftime('%Y-%m-%d %H:%M')
                        parsed = pd.to_datetime(x, errors='coerce')
                        if pd.notna(parsed):
                            return parsed.strftime('%Y-%m-%d %H:%M')
                        return ''
                    except Exception as e:
                        logger.warning(f"Could not format datetime {x}: {e}")
                        return str(x) if x else ''
                
                display_df[col] = display_df[col].apply(format_datetime)
                logger.info(f"After formatting {col}: {display_df[col].head(3).tolist()}")
        
        # Format date-only columns
        date_cols = ['Target Date']
        for col in date_cols:
            if col in display_df.columns:
                def format_date(x):
                    if pd.isna(x) or x in ['', None, 'None', 'NaT']:
                        return ''
                    try:
                        parsed = pd.to_datetime(x, errors='coerce')
                        if pd.notna(parsed):
                            return parsed.strftime('%Y-%m-%d')
                        return ''
                    except:
                        return str(x) if x else ''
                
                display_df[col] = display_df[col].apply(format_date)
        
        # Write data to sheet
        for r_idx, row in enumerate(display_df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                # Handle any remaining NaT values
                if pd.isna(value) or value in ['NaT', 'None', None]:
                    value = ''
                
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Color code status column if it exists
                if 'Status' in display_df.columns:
                    status_col_idx = list(display_df.columns).index('Status') + 1
                    if c_idx == status_col_idx and value:
                        color = self.processor.get_status_color(value)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                # Make notes column wider and wrap text
                col_name = display_df.columns[c_idx - 1] if c_idx <= len(display_df.columns) else None
                if col_name == 'Work History & Notes':
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Write headers
        for c_idx, col_name in enumerate(display_df.columns, start=1):
            ws.cell(row=1, column=c_idx, value=col_name)
            
            # Set specific column widths
            if col_name == 'Work History & Notes':
                ws.column_dimensions[chr(64 + c_idx)].width = 60
            elif col_name in ['Title', 'Component']:
                ws.column_dimensions[chr(64 + c_idx)].width = 25
            elif col_name in ['Created', 'Last Activity']:
                ws.column_dimensions[chr(64 + c_idx)].width = 18
        
        self.styler.style_header_row(ws, row=1)
        self.styler.auto_adjust_column_width(ws)
        
        # Set row height to auto-fit wrapped text
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = None  # Auto-height
    
    def _create_by_trade_sheet(self, wb, defects_df):
        """Create defects by trade sheet"""
        ws = wb.create_sheet("🔧 By Trade")
        
        if defects_df.empty:
            ws['A1'] = 'No defect data available'
            return
        
        trade_summary = defects_df.groupby('trade').agg({
            'id': 'count',
            'urgency': lambda x: (x == 'Urgent').sum()
        }).reset_index()
        
        trade_summary.columns = ['Trade', 'Total Defects', 'Urgent Count']
        trade_summary = trade_summary.sort_values('Total Defects', ascending=False)
        
        for r_idx, row in enumerate(trade_summary.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        for c_idx, col_name in enumerate(trade_summary.columns, start=1):
            ws.cell(row=1, column=c_idx, value=col_name)
        
        self.styler.style_header_row(ws, row=1)
        self.styler.auto_adjust_column_width(ws)
        
        try:
            self._add_bar_chart(ws, 'Defects by Trade', 'Trade', 'Number of Defects')
        except:
            pass
    
    def _create_by_unit_sheet(self, wb, defects_df):
        """Create defects by unit sheet"""
        ws = wb.create_sheet("🏠 By Unit")
        
        if defects_df.empty:
            ws['A1'] = 'No defect data available'
            return
        
        unit_summary = defects_df.groupby('unit').size().reset_index(name='Total Defects')
        unit_summary = unit_summary.sort_values('Total Defects', ascending=False).head(20)
        
        for r_idx, row in enumerate(unit_summary.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        for c_idx, col_name in enumerate(unit_summary.columns, start=1):
            ws.cell(row=1, column=c_idx, value=col_name)
        
        self.styler.style_header_row(ws, row=1)
        self.styler.auto_adjust_column_width(ws)
    
    def _create_progress_sheet(self, wb, work_orders_df):
        """Create progress tracking sheet"""
        ws = wb.create_sheet("📈 Progress")
        
        # ✅ FIXED: Group by actual status values, count work orders only
        status_summary = work_orders_df.groupby('status').agg({
            'id': 'count'
        }).reset_index()
        
        status_summary.columns = ['Status', 'Work Orders']
        
        # Sort by logical order
        status_order = {'pending': 1, 'in_progress': 2, 'waiting_approval': 3, 'approved': 4}
        status_summary['sort_order'] = status_summary['Status'].map(status_order)
        status_summary = status_summary.sort_values('sort_order').drop('sort_order', axis=1)
        
        for r_idx, row in enumerate(status_summary.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Color code status
                if c_idx == 1:  # Status column
                    color = self.processor.get_status_color(value)
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        for c_idx, col_name in enumerate(status_summary.columns, start=1):
            ws.cell(row=1, column=c_idx, value=col_name)
        
        self.styler.style_header_row(ws, row=1)
        self.styler.auto_adjust_column_width(ws)
    
    def _create_photos_sheet(self, wb, defects_df):
        """Create photo references sheet with embedded images"""
        ws = wb.create_sheet("📷 Photos")
        
        # Get all work order files from database
        conn = None
        try:
            conn = self._get_fresh_connection()
            
            # Try to get ONLY IMAGE files from work_order_files table
            try:
                photos_df = pd.read_sql_query("""
                    SELECT 
                        wof.work_order_id,
                        wof.original_filename,
                        wof.file_path,
                        wof.file_type,
                        wof.uploaded_at,
                        wo.unit,
                        wo.room,
                        wo.component,
                        wo.trade,
                        wo.status
                    FROM work_order_files wof
                    LEFT JOIN inspector_work_orders wo ON wof.work_order_id = wo.id
                    WHERE wof.file_type LIKE 'image%'
                    ORDER BY wof.uploaded_at DESC
                """, conn)
                logger.info(f"Found {len(photos_df)} image files")
            except:
                # Try alternative table name
                try:
                    photos_df = pd.read_sql_query("""
                        SELECT 
                            wof.work_order_id,
                            wof.original_filename,
                            wof.file_path,
                            wof.file_type,
                            wof.uploaded_at
                        FROM inspector_work_order_files wof
                        WHERE wof.file_type LIKE 'image%'
                        ORDER BY wof.uploaded_at DESC
                    """, conn)
                except Exception as e:
                    logger.error(f"Could not query photos: {e}")
                    photos_df = pd.DataFrame()
            
            if photos_df.empty:
                ws['A1'] = '📷 No photos found in the system'
                ws['A1'].font = Font(bold=True, size=12)
                ws['A3'] = 'Photos will appear here once builders upload them with their work orders.'
                ws['A4'] = 'Note: This sheet only shows image files (JPG, PNG, etc.)'
                return []
            
            # Create header
            ws['A1'] = '📷 PHOTO REFERENCES'
            ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color=self.styler.HEADER_COLOR, 
                                       end_color=self.styler.HEADER_COLOR, 
                                       fill_type="solid")
            ws.merge_cells('A1:G1')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 25  # Title row height
            
            # Column headers (row 2)
            headers = ['Preview', 'Work Order', 'Filename', 'Unit', 'Room', 'Component', 'Trade']
            for c_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=c_idx, value=header)
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color=self.styler.HEADER_COLOR, 
                                       end_color=self.styler.HEADER_COLOR, 
                                       fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.row_dimensions[2].height = 20  # Header row height
            
            # Set column widths
            ws.column_dimensions['A'].width = 25  # Preview column (wide for images)
            ws.column_dimensions['B'].width = 12  # WO ID
            ws.column_dimensions['C'].width = 30  # Filename
            ws.column_dimensions['D'].width = 10  # Unit
            ws.column_dimensions['E'].width = 15  # Room
            ws.column_dimensions['F'].width = 20  # Component
            ws.column_dimensions['G'].width = 20  # Trade
            
            # Add data rows with images
            current_row = 3  # Start at row 3 (after title row 1 and header row 2)
            images_added = 0
            images_failed = 0
            temp_files = []  # Track temp files to clean up later
            
            for idx, row in photos_df.iterrows():
                file_path = row.get('file_path', '')
                
                # Determine row height based on whether it's an image
                is_image = file_path.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')) if file_path else False
                
                if is_image:
                    # Set taller row height for images (in points, not pixels)
                    ws.row_dimensions[current_row].height = 80
                else:
                    # Normal row height for non-images
                    ws.row_dimensions[current_row].height = 20
                
                # Add text data
                ws.cell(row=current_row, column=2, value=row.get('work_order_id', ''))
                ws.cell(row=current_row, column=3, value=row.get('original_filename', ''))
                ws.cell(row=current_row, column=4, value=row.get('unit', ''))
                ws.cell(row=current_row, column=5, value=row.get('room', ''))
                ws.cell(row=current_row, column=6, value=row.get('component', ''))
                ws.cell(row=current_row, column=7, value=row.get('trade', ''))
                
                # Try to embed the image or show info for non-images
                if file_path and os.path.exists(file_path):
                    try:
                        # Check if it's actually an image file
                        if file_path.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                            # Load image and get original dimensions
                            img = PILImage.open(file_path)
                            original_width = img.width
                            original_height = img.height
                            
                            # Target dimensions to fit nicely in Excel cell
                            max_height_pixels = 100  # Fit within row with padding
                            max_width_pixels = 140   # Fit within column with padding
                            
                            # Calculate scale to fit within constraints (maintain aspect ratio)
                            scale_height = max_height_pixels / original_height if original_height > max_height_pixels else 1
                            scale_width = max_width_pixels / original_width if original_width > max_width_pixels else 1
                            scale = min(scale_height, scale_width)
                            
                            # Always resize to ensure consistent sizing
                            new_width = int(original_width * scale)
                            new_height = int(original_height * scale)
                            img = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
                            
                            # Use proper temporary directory
                            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                                temp_path = tmp.name
                                img.save(temp_path, 'PNG')
                            
                            # Track temp file for cleanup after workbook is saved
                            temp_files.append(temp_path)
                            
                            # Add to Excel with proper positioning
                            xl_img = XLImage(temp_path)
                            xl_img.anchor = f'A{current_row}'
                            ws.add_image(xl_img)
                            
                            images_added += 1
                            logger.info(f"✅ Embedded image: {file_path} (resized to {new_width}x{new_height})")
                        else:
                            # Should not happen since we filter for images only in SQL
                            cell = ws.cell(row=current_row, column=1, value="[Not an image]")
                            cell.font = Font(color="999999", italic=True)
                            logger.warning(f"Non-image file in photos sheet: {file_path}")
                            
                    except Exception as e:
                        logger.warning(f"Could not embed image {file_path}: {e}")
                        cell = ws.cell(row=current_row, column=1, value=f"[Error loading]")
                        cell.font = Font(color="FF6600")
                        images_failed += 1
                else:
                    cell = ws.cell(row=current_row, column=1, value="[File not found]")
                    cell.font = Font(color="FF0000")
                    logger.warning(f"File not found: {file_path}")
                    images_failed += 1
                
                current_row += 1
            
            # Update summary header with counts
            summary_text = f'📷 PHOTO REFERENCES - {len(photos_df)} files | {images_added} embedded | {images_failed} failed'
            ws['A1'] = summary_text
            ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="4472C4", 
                                       end_color="4472C4", 
                                       fill_type="solid")
            ws.merge_cells('A1:G1')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            logger.info(f"✅ Photos sheet created: {images_added} embedded, {images_failed} failed")
            
            # Return list of temp files to be cleaned up after workbook is saved
            return temp_files
            
        except Exception as e:
            logger.error(f"Error creating photos sheet: {e}")
            import traceback
            logger.error(traceback.format_exc())
            ws['A1'] = 'Error loading photo information'
            ws['A1'].font = Font(bold=True, size=12, color="FF0000")
            ws['A3'] = f'Error: {str(e)}'
            return []  # Return empty list on error
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass
    
    def _create_files_sheet(self, wb, defects_df):
        """Create a sheet listing all non-image file attachments"""
        ws = wb.create_sheet("📎 Files")
        
        conn = None
        try:
            conn = self._get_fresh_connection()
            
            # Get all NON-IMAGE files from work_order_files table
            try:
                # First, try with all columns including file_size
                files_df = pd.read_sql_query("""
                    SELECT 
                        wof.work_order_id,
                        wof.original_filename,
                        wof.file_path,
                        wof.file_type,
                        wof.uploaded_at,
                        wo.unit,
                        wo.room,
                        wo.component,
                        wo.trade,
                        wo.status
                    FROM work_order_files wof
                    LEFT JOIN inspector_work_orders wo ON wof.work_order_id = wo.id
                    WHERE wof.file_type NOT LIKE 'image%'
                    ORDER BY wof.uploaded_at DESC
                """, conn)
                logger.info(f"Found {len(files_df)} non-image files")
            except Exception as e1:
                logger.warning(f"First query failed: {e1}, trying alternative...")
                # Try alternative table name
                try:
                    files_df = pd.read_sql_query("""
                        SELECT 
                            wof.work_order_id,
                            wof.original_filename,
                            wof.file_path,
                            wof.file_type,
                            wof.uploaded_at
                        FROM inspector_work_order_files wof
                        WHERE wof.file_type NOT LIKE 'image%'
                        ORDER BY wof.uploaded_at DESC
                    """, conn)
                    logger.info(f"Alternative query: Found {len(files_df)} non-image files")
                except Exception as e2:
                    logger.error(f"Both queries failed: {e2}")
                    # Last resort: get ALL files and filter in Python
                    try:
                        files_df = pd.read_sql_query("""
                            SELECT 
                                wof.work_order_id,
                                wof.original_filename,
                                wof.file_path,
                                wof.file_type,
                                wof.uploaded_at,
                                wo.unit,
                                wo.room,
                                wo.component,
                                wo.trade,
                                wo.status
                            FROM work_order_files wof
                            LEFT JOIN inspector_work_orders wo ON wof.work_order_id = wo.id
                            ORDER BY wof.uploaded_at DESC
                        """, conn)
                        # Filter out images in Python
                        files_df = files_df[~files_df['file_type'].str.contains('image', case=False, na=False)]
                        logger.info(f"Fallback query: Found {len(files_df)} non-image files after filtering")
                    except Exception as e3:
                        logger.error(f"All queries failed: {e3}")
                        files_df = pd.DataFrame()
            
            # DEBUG: Show what we found
            if not files_df.empty:
                logger.info(f"Files found - Types: {files_df['file_type'].unique()}")
                logger.info(f"Sample files:\n{files_df[['original_filename', 'file_type']].head()}")
            else:
                logger.warning("No files found in database!")
            
            if files_df.empty:
                ws['A1'] = '📎 No document files found'
                ws['A1'].font = Font(bold=True, size=12)
                ws['A3'] = 'Document files (PDFs, videos, etc.) will appear here when uploaded.'
                ws['A4'] = 'Note: Only non-image files are shown here. Check the Photos sheet for images.'
                return
            
            # Create header
            ws['A1'] = f'📎 DOCUMENT FILES - {len(files_df)} files'
            ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color=self.styler.HEADER_COLOR, 
                                       end_color=self.styler.HEADER_COLOR, 
                                       fill_type="solid")
            ws.merge_cells('A1:J1')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 25
            
            # Column headers (row 2)
            headers = ['Work Order', 'Filename', 'Type', 'Size', 'Uploaded', 'Unit', 'Room', 'Component', 'Trade', 'File Path']
            for c_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=c_idx, value=header)
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color=self.styler.HEADER_COLOR, 
                                       end_color=self.styler.HEADER_COLOR, 
                                       fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.row_dimensions[2].height = 20
            
            # Set column widths
            ws.column_dimensions['A'].width = 12  # WO ID
            ws.column_dimensions['B'].width = 35  # Filename
            ws.column_dimensions['C'].width = 25  # Type
            ws.column_dimensions['D'].width = 10  # Size
            ws.column_dimensions['E'].width = 18  # Uploaded
            ws.column_dimensions['F'].width = 10  # Unit
            ws.column_dimensions['G'].width = 15  # Room
            ws.column_dimensions['H'].width = 20  # Component
            ws.column_dimensions['I'].width = 20  # Trade
            ws.column_dimensions['J'].width = 60  # File Path
            
            # Add data rows
            current_row = 3
            for idx, row in files_df.iterrows():
                ws.cell(row=current_row, column=1, value=row.get('work_order_id', ''))
                ws.cell(row=current_row, column=2, value=row.get('original_filename', ''))
                
                # Format file type nicely
                file_type = row.get('file_type', '')
                if 'pdf' in file_type.lower():
                    file_type_display = '📄 PDF Document'
                elif 'video' in file_type.lower():
                    file_type_display = '🎥 Video'
                elif 'audio' in file_type.lower():
                    file_type_display = '🔊 Audio'
                elif 'excel' in file_type.lower() or 'spreadsheet' in file_type.lower():
                    file_type_display = '📊 Spreadsheet'
                elif 'word' in file_type.lower() or 'document' in file_type.lower():
                    file_type_display = '📝 Document'
                else:
                    file_type_display = file_type
                ws.cell(row=current_row, column=3, value=file_type_display)
                
                # Format file size - check if column exists
                if 'file_size' in files_df.columns:
                    file_size = row.get('file_size', 0)
                    if file_size:
                        if file_size < 1024:
                            size_display = f"{file_size} B"
                        elif file_size < 1024 * 1024:
                            size_display = f"{file_size / 1024:.1f} KB"
                        else:
                            size_display = f"{file_size / (1024 * 1024):.1f} MB"
                    else:
                        size_display = "—"
                else:
                    # Try to get file size from file path
                    file_path = row.get('file_path', '')
                    if file_path and os.path.exists(file_path):
                        try:
                            file_size = os.path.getsize(file_path)
                            if file_size < 1024:
                                size_display = f"{file_size} B"
                            elif file_size < 1024 * 1024:
                                size_display = f"{file_size / 1024:.1f} KB"
                            else:
                                size_display = f"{file_size / (1024 * 1024):.1f} MB"
                        except:
                            size_display = "—"
                    else:
                        size_display = "—"
                
                ws.cell(row=current_row, column=4, value=size_display)
                
                # Format uploaded date
                uploaded = row.get('uploaded_at', '')
                if uploaded:
                    try:
                        uploaded_dt = pd.to_datetime(uploaded)
                        uploaded_display = uploaded_dt.strftime('%Y-%m-%d %H:%M')
                    except:
                        uploaded_display = str(uploaded)
                else:
                    uploaded_display = ""
                ws.cell(row=current_row, column=5, value=uploaded_display)
                
                ws.cell(row=current_row, column=6, value=row.get('unit', ''))
                ws.cell(row=current_row, column=7, value=row.get('room', ''))
                ws.cell(row=current_row, column=8, value=row.get('component', ''))
                ws.cell(row=current_row, column=9, value=row.get('trade', ''))
                
                # Add file path in the last column
                file_path = row.get('file_path', '')
                path_cell = ws.cell(row=current_row, column=10, value=file_path)
                path_cell.font = Font(size=9, color="0000FF")  # Blue, smaller font
                
                current_row += 1
            
            logger.info(f"✅ Files sheet created with {len(files_df)} documents")
            
        except Exception as e:
            logger.error(f"Error creating files sheet: {e}")
            ws['A1'] = 'Error loading file information'
            ws['A1'].font = Font(bold=True, size=12, color="FF0000")
            ws['A3'] = f'Error: {str(e)}'
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass
    
    def _add_bar_chart(self, ws, title, x_title, y_title):
        """Add a bar chart to worksheet"""
        try:
            chart = BarChart()
            chart.title = title
            chart.x_axis.title = x_title
            chart.y_axis.title = y_title
            chart.height = 10
            chart.width = 20
            
            data = Reference(ws, min_col=2, min_row=1, max_row=min(ws.max_row, 15))
            cats = Reference(ws, min_col=1, min_row=2, max_row=min(ws.max_row, 15))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, "E2")
        except Exception as e:
            logger.warning(f"Could not add chart: {e}")
    
    def generate_report_package(self, builder_name=None, inspection_id=None, include_photos=False, status_filter=None):
        """Generate a ZIP package containing the Excel report and all attached files"""
        import zipfile
        import shutil
        
        temp_files = []
        temp_dir = None
        
        try:
            # Generate the Excel report WITH Files sheet
            excel_output = self.generate_excel_report(
                builder_name=builder_name,
                inspection_id=inspection_id,
                include_photos=include_photos,
                status_filter=status_filter,
                include_files_sheet=True  # Include Files sheet in ZIP package
            )
            
            if not excel_output:
                return None
            
            # Create temporary directory for the package
            temp_dir = tempfile.mkdtemp()
            attachments_dir = os.path.join(temp_dir, 'attachments')
            os.makedirs(attachments_dir, exist_ok=True)
            
            # Save Excel file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            excel_filename = f'Work_Orders_Report_{timestamp}.xlsx'
            excel_path = os.path.join(temp_dir, excel_filename)
            with open(excel_path, 'wb') as f:
                f.write(excel_output.getvalue())
            
            # Get all files from database
            conn = self._get_fresh_connection()
            files_copied = 0
            files_failed = 0
            
            try:
                query = """
                    SELECT 
                        wof.work_order_id,
                        wof.original_filename,
                        wof.file_path,
                        wof.file_type
                    FROM work_order_files wof
                    INNER JOIN inspector_work_orders wo ON wof.work_order_id = wo.id
                    WHERE 1=1
                """
                
                params = []
                if inspection_id:
                    query += " AND wo.inspection_id = ?"
                    params.append(inspection_id)
                
                if status_filter:
                    query += f" AND {status_filter}"
                
                files_df = pd.read_sql_query(query, conn, params=params if params else None)
                logger.info(f"Found {len(files_df)} files to copy")
                
                # Copy files to attachments folder
                for idx, row in files_df.iterrows():
                    source_path = row['file_path']
                    if os.path.exists(source_path):
                        # Create work order subfolder
                        wo_folder = os.path.join(attachments_dir, str(row['work_order_id']))
                        os.makedirs(wo_folder, exist_ok=True)
                        
                        # Copy file with original filename
                        dest_path = os.path.join(wo_folder, row['original_filename'])
                        try:
                            shutil.copy2(source_path, dest_path)
                            files_copied += 1
                            logger.info(f"✅ Copied: {row['original_filename']}")
                        except Exception as e:
                            logger.warning(f"❌ Could not copy {source_path}: {e}")
                            files_failed += 1
                    else:
                        logger.warning(f"File not found: {source_path}")
                        files_failed += 1
                
                logger.info(f"Copied {files_copied} files, {files_failed} failed")
                
            finally:
                conn.close()
            
            # Create README file with instructions
            readme_path = os.path.join(temp_dir, 'README.txt')
            with open(readme_path, 'w', encoding='utf-8') as f:
                f.write(f"""Work Orders Report Package
{'=' * 50}

📦 Package Contents:
  • {excel_filename}: Main Excel report with all work order data
  • attachments/: Folder containing all uploaded files organized by work order
  • README.txt: This file

📊 How to Use the Report:
  1. Open {excel_filename} to view the complete report
  2. The report contains multiple sheets:
     - 📊 Executive Summary: Overview statistics
     - 📋 Work Orders: Detailed work order list
     - 🔧 By Trade: Defects grouped by trade
     - 🏠 By Unit: Defects grouped by unit
     - 📈 Progress: Status summary
     - 📷 Photos: Embedded image thumbnails
     - 📎 Files: Reference list of all documents

📁 File Organization:
  The attachments folder is organized as:
  
  attachments/
    ├── [work-order-id-1]/
    │   ├── photo1.jpg
    │   ├── document.pdf
    │   └── ...
    ├── [work-order-id-2]/
    │   ├── photo2.jpg
    │   └── ...
    └── ...

💡 Tips:
  • Photos are embedded directly in the Photos sheet - no need to open files
  • Use the Files sheet to see which documents are available
  • File paths in the Files sheet show: ./attachments/[work-order-id]/[filename]
  • Keep this entire folder structure together for the paths to work

📊 Report Statistics:
  • Files included: {files_copied}
  • Files not found: {files_failed}
  • Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

---
Building Inspection System V3
""")
            
            # Create ZIP file
            zip_output = BytesIO()
            with zipfile.ZipFile(zip_output, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # Add Excel report
                zipf.write(excel_path, excel_filename)
                
                # Add README
                zipf.write(readme_path, 'README.txt')
                
                # Add all files from attachments folder
                if os.path.exists(attachments_dir) and os.listdir(attachments_dir):
                    for root, dirs, files in os.walk(attachments_dir):
                        for file in files:
                            file_path = os.path.join(root, file)
                            arcname = os.path.relpath(file_path, temp_dir)
                            zipf.write(file_path, arcname)
            
            zip_output.seek(0)
            logger.info(f"✅ ZIP package created: {files_copied} files included")
            return zip_output
            
        except Exception as e:
            logger.error(f"❌ Error creating ZIP package: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None
            
        finally:
            # Clean up temporary directory
            if temp_dir and os.path.exists(temp_dir):
                try:
                    shutil.rmtree(temp_dir)
                    logger.info("🧹 Cleaned up temp directory")
                except Exception as e:
                    logger.warning(f"Could not clean up temp dir: {e}")


def add_builder_report_ui(db_manager):
    """Add report generation UI"""
    
    st.info("💡 Generate comprehensive reports of work orders and defects")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        conn = None
        try:
            # Use fresh connection
            db_path = db_manager.db_path if hasattr(db_manager, 'db_path') else "building_inspection.db"
            conn = sqlite3.connect(db_path, check_same_thread=False, detect_types=0)
            
            # Get inspections with work order counts by status
            inspections_df = pd.read_sql_query("""
                SELECT 
                    i.id, 
                    i.inspection_date, 
                    b.name as building_name,
                    COUNT(CASE WHEN wo.status = 'pending' THEN 1 END) as pending,
                    COUNT(CASE WHEN wo.status = 'in_progress' THEN 1 END) as in_progress,
                    COUNT(CASE WHEN wo.status = 'waiting_approval' THEN 1 END) as waiting,
                    COUNT(CASE WHEN wo.status = 'approved' THEN 1 END) as approved,
                    COUNT(*) as total
                FROM inspector_inspections i
                JOIN inspector_buildings b ON i.building_id = b.id
                LEFT JOIN inspector_work_orders wo ON i.id = wo.inspection_id
                GROUP BY i.id, i.inspection_date, b.name
                HAVING total > 0
                ORDER BY i.inspection_date DESC
                LIMIT 20
            """, conn)
            
            logger.info(f"Found {len(inspections_df)} inspections with work orders")
            
        except Exception as e:
            logger.error(f"Error loading inspections: {e}")
            inspections_df = pd.DataFrame()
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass
        
        # Build inspection options
        inspection_options = ["All Inspections"]
        inspection_id_map = {0: None}  # Map index to inspection ID
        
        if not inspections_df.empty:
            for idx, row in inspections_df.iterrows():
                option_text = f"{row['building_name']} - {row['inspection_date']} (P:{row['pending']} A:{row['in_progress']+row['waiting']+row['approved']})"
                inspection_options.append(option_text)
                inspection_id_map[len(inspection_options) - 1] = row['id']
        
        selected_inspection_idx = st.selectbox(
            "📋 Select Inspection", 
            range(len(inspection_options)),
            format_func=lambda x: inspection_options[x],
            key="report_inspection_selector",
            help="Numbers show (P:Pending A:Active+Waiting+Approved)"
        )
    
    with col2:
        # Status filter option
        status_filter_option = st.selectbox(
            "📊 Status Filter",
            ["All Statuses", "Active Only (Exclude Pending)", "Non-Pending Only", "Pending Only"],
            key="report_status_filter",
            help="Filter which work orders to include"
        )
    
    # Put photos checkbox on a new row for better layout
    include_photos = st.checkbox(
        "📷 Include Photos Sheet", 
        value=False, 
        help="Include a sheet with embedded image thumbnails",
        key="report_include_photos"
    )
    
    st.write("---")
    st.write("**Choose report format:**")
    
    col_btn1, col_btn2 = st.columns(2)
    
    with col_btn1:
        excel_only_button = st.button(
            "📊 Excel Report Only",
            help="Download Excel report without attachments",
            use_container_width=True,
            key="report_excel_only_button"
        )
    
    with col_btn2:
        zip_package_button = st.button(
            "📦 ZIP Package (Report + Files)", 
            type="primary",
            help="Download ZIP containing Excel report and all uploaded files",
            use_container_width=True,
            key="report_zip_package_button"
        )
    
    # Handle Excel Only download
    if excel_only_button:
        with st.spinner("🔄 Generating Excel report..."):
            try:
                # Get the actual inspection ID from the map
                inspection_id = inspection_id_map.get(selected_inspection_idx)
                
                # Map status filter to SQL condition
                status_sql = None
                if status_filter_option == "Active Only (Exclude Pending)":
                    status_sql = "wo.status != 'pending'"
                elif status_filter_option == "Non-Pending Only":
                    status_sql = "wo.status IN ('in_progress', 'waiting_approval', 'approved')"
                elif status_filter_option == "Pending Only":
                    status_sql = "wo.status = 'pending'"
                
                # Debug output with actual values
                selected_text = inspection_options[selected_inspection_idx]
                logger.info(f"🔍 Excel Report Generation:")
                logger.info(f"   - Selected: {selected_text}")
                logger.info(f"   - Inspection ID: {inspection_id}")
                logger.info(f"   - Include Photos: {include_photos}")
                logger.info(f"   - Status Filter: {status_filter_option}")
                
                st.info(f"🔍 Generating Excel report for: **{selected_text}**")
                
                generator = BuilderReportGenerator(db_manager)
                excel_file = generator.generate_excel_report(
                    builder_name=None,
                    inspection_id=inspection_id,
                    include_photos=include_photos,
                    status_filter=status_sql,
                    include_files_sheet=False  # No Files sheet in standalone Excel
                )
                
                if excel_file:
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    inspection_name = selected_text.replace(' - ', '_').replace('(', '').replace(')', '').replace(':', '')
                    if len(inspection_name) > 50:
                        inspection_name = inspection_name[:50]
                    
                    filename = f"Work_Orders_{inspection_name}_{timestamp}.xlsx"
                    
                    st.download_button(
                        label="📥 Download Excel Report",
                        data=excel_file,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="report_download_excel_button"
                    )
                    
                    st.success("✅ Excel report generated!")
                    
                    with st.expander("📊 Report Summary"):
                        st.write(f"**Format:** Excel Only (no attachments)")
                        st.write(f"**Inspection:** {selected_text}")
                        st.write(f"**Photos:** {'Embedded in report' if include_photos else 'Not included'}")
                else:
                    st.warning("⚠️ No data available")
                    
            except Exception as e:
                logger.error(f"Excel report error: {e}")
                st.error(f"❌ Error: {e}")
    
    # Handle ZIP Package download
    if zip_package_button:
        with st.spinner("🔄 Generating ZIP package with all files..."):
            try:
                # Get the actual inspection ID from the map
                inspection_id = inspection_id_map.get(selected_inspection_idx)
                
                # Map status filter to SQL condition
                status_sql = None
                if status_filter_option == "Active Only (Exclude Pending)":
                    status_sql = "wo.status != 'pending'"
                elif status_filter_option == "Non-Pending Only":
                    status_sql = "wo.status IN ('in_progress', 'waiting_approval', 'approved')"
                elif status_filter_option == "Pending Only":
                    status_sql = "wo.status = 'pending'"
                
                selected_text = inspection_options[selected_inspection_idx]
                logger.info(f"🔍 ZIP Package Generation:")
                logger.info(f"   - Selected: {selected_text}")
                logger.info(f"   - Inspection ID: {inspection_id}")
                logger.info(f"   - Include Photos: {include_photos}")
                
                st.info(f"📦 Creating package for: **{selected_text}** (this may take a moment...)")
                
                generator = BuilderReportGenerator(db_manager)
                zip_file = generator.generate_report_package(
                    builder_name=None,
                    inspection_id=inspection_id,
                    include_photos=include_photos,
                    status_filter=status_sql
                )
                
                if zip_file:
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    inspection_name = selected_text.replace(' - ', '_').replace('(', '').replace(')', '').replace(':', '')
                    if len(inspection_name) > 50:
                        inspection_name = inspection_name[:50]
                    
                    filename = f"Work_Orders_Package_{inspection_name}_{timestamp}.zip"
                    
                    st.download_button(
                        label="📦 Download ZIP Package",
                        data=zip_file,
                        file_name=filename,
                        mime="application/zip",
                        use_container_width=True,
                        key="report_download_zip_button"
                    )
                    
                    st.success("✅ ZIP package created successfully!")
                    
                    with st.expander("📦 Package Contents", expanded=True):
                        st.write(f"**Format:** ZIP Package")
                        st.write(f"**Inspection:** {selected_text}")
                        st.write(f"**Includes:**")
                        st.write(f"  • Excel report with all sheets")
                        st.write(f"  • {'📷 Embedded photos' if include_photos else 'No photos'}")
                        st.write(f"  • 📎 Files sheet (reference list)")
                        st.write(f"  • 📁 attachments/ folder with all uploaded files")
                        st.write(f"  • 📄 README.txt with instructions")
                        st.info("💡 **Tip:** Extract the ZIP file and keep all files together. The Files sheet shows relative paths that work within the extracted folder.")
                else:
                    st.warning("⚠️ No data available")
                    
            except Exception as e:
                logger.error(f"ZIP package error: {e}")
                st.error(f"❌ Error: {e}")
                import traceback
                with st.expander("🔍 Error Details"):
                    st.code(traceback.format_exc())